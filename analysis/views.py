from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from collections import defaultdict, OrderedDict
from types import SimpleNamespace
from django.http import JsonResponse, FileResponse, HttpResponse, Http404
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from PyPDF2 import PdfReader, PdfWriter 
from openpyxl import load_workbook  
import csv
from django.utils import timezone
from datetime import timedelta
from django.views.decorators.http import require_POST
from django.db.models import IntegerField, Value, Count
from django.db.models.functions import Cast, Coalesce, TruncMonth
from django.db.models import Q
from datetime import date
from django.db.models import Avg, Count, Q, Case, When, IntegerField
import re
import zipfile
from io import BytesIO
from reportlab.pdfgen import canvas
import json
from django.contrib.staticfiles.storage import staticfiles_storage
import copy
import numpy as np
import pytesseract
from pytesseract import Output
from rapidfuzz import fuzz, process
from PIL import Image, ImageOps
from django.urls import reverse
from django.shortcuts import redirect
import subprocess
from django.core import management
from io import StringIO
from django.db import transaction

from analysis.models import (
    Teacher,
    Student,
    Grade,
    Section,
    Subject,
    AnswerKey,
    ScanResult,
    Institution,
    InstitutionAdmin,
    TeacherClassAssignment,
    SchoolYear,
    GradingPeriod,
)



from .forms import (
    AddUserForm,
    EditUserForm,
    GradeForm,
    SectionForm,
    SubjectForm,
    StudentForm,  
    InstitutionForm, 
    InstitutionAdminForm, 
    UserAdminForm,
    SchoolYear,
    GradingPeriodForm,
)

from .utils import (
    read_word_answer_key,
    generate_bubble_sheet_pdf,
    read_answer_key_file,
)

import os
import re
import cv2
import uuid
import base64
from django.db import IntegrityError
from django.utils import timezone
from datetime import datetime
from pathlib import Path
import json
import sys

from .omr_knn.omr_knn_pipeline import (
    load_knn_model,
    analyze_sheet,
    score_sheet,
    annotate_sheet,
)


# Lazy-loaded KNN model (prevents Django crash on import)
KNN_MODEL = None
KNN_CLASS_NAMES = None
KNN_IMG_SIZE = None
KNN_CAL = None

def get_knn():
    global KNN_MODEL, KNN_CLASS_NAMES, KNN_IMG_SIZE, KNN_CAL
    if KNN_MODEL is None:
        out = load_knn_model()

        # Supports BOTH return styles: (3) or (4)
        if len(out) == 3:
            KNN_MODEL, KNN_CLASS_NAMES, KNN_IMG_SIZE = out
            KNN_CAL = {}
        else:
            KNN_MODEL, KNN_CLASS_NAMES, KNN_IMG_SIZE, KNN_CAL = out

    return KNN_MODEL, KNN_CLASS_NAMES, KNN_IMG_SIZE, KNN_CAL

# Super Admin Login
def super_admin_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None and user.is_superuser:
            login(request, user)
            return redirect('institution_admin_dashboard')  # Redirect to admin dashboard after login
        else:
            messages.error(request, "Invalid username or password, or you are not authorized.")
            return redirect('super_admin_login') 

    return render(request, "institution_admin/login.html")  # Correct template path

# Super Admin Logout
def super_admin_logout(request):
    logout(request)
    return redirect('super_admin_login')  # Redirect to login after logout

# Admin Login for Institution Admin
def institution_admin_login(request):
    """Login for Institution Admin"""
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        # Authenticate the user
        user = authenticate(request, username=username, password=password)

        if user is not None:
            try:
                institution_admin = InstitutionAdmin.objects.get(user=user)
                login(request, user)  # Log the user in if they are an institution admin
                return redirect("institution_dashboard", institution_id=institution_admin.institution.id)
            except InstitutionAdmin.DoesNotExist:
                messages.error(request, "You are not authorized to access this page.")
                return redirect("institution_admin_login")
        else:
            messages.error(request, "Invalid username or password.")
            return redirect("institution_admin_login")

    return render(request, "institution_admin/admin_login.html")  # Correct template path

@login_required
def institution_admin_dashboard(request):
    if not request.user.is_superuser:
        return HttpResponse("Unauthorized", status=403)

    total_admins = User.objects.filter(is_staff=True, is_superuser=False).count()
    total_teachers = User.objects.filter(is_staff=False, is_superuser=False).count()
    
    recent_logs = []
    recent_logs_count = 0

    context = {
        "total_institutions": Institution.objects.count(),
        "total_institution_admins": InstitutionAdmin.objects.count(),
        "total_users": User.objects.count(),

        "total_admin_users": total_admins,
        "total_teacher_users": total_teachers,

        "recent_logs": recent_logs,
        "recent_logs_count": recent_logs_count if isinstance(recent_logs, list) else recent_logs.count(),
    }

    return render(request, "institution_admin/admin_dashboard.html", context)

# Add Institution (Only accessible by superuser)
@login_required
def add_institution(request, institution_id=None):
    # Ensure only superusers can access this view
    if not request.user.is_superuser:
        messages.error(request, "You are not allowed to access this page.")
        return redirect("admin_dashboard")  # Adjust this according to your dashboard name

    editing = None
    if institution_id is not None:
        editing = get_object_or_404(Institution, id=institution_id)

    if request.method == "POST":
        form = InstitutionForm(request.POST, instance=editing)
        if form.is_valid():
            inst = form.save()
            if editing:
                messages.success(request, f"Institution updated: {inst.name}")
            else:
                messages.success(request, f"Institution created: {inst.name}")
            # Redirect to institution detail page after successful save
            return redirect('institution_detail', institution_id=inst.id)  # Adjusted to use the institution's ID for detail page
        else:
            messages.error(request, "Please fix the errors below.")
    else:
        form = InstitutionForm(instance=editing)

    institutions = Institution.objects.all().order_by("-id")

    return render(
        request,
        "institution_admin/add_institution.html",
        {
            "form": form,
            "institutions": institutions,
            "editing": editing,
        },
    )

@login_required
def institution_detail(request, institution_id):
    institution = get_object_or_404(Institution, id=institution_id)

    # Ensure only superusers can access the detail page
    if not request.user.is_superuser:
        messages.error(request, "You are not allowed to access this page.")
        return redirect("admin_dashboard")

    institution_admins = (
        InstitutionAdmin.objects
        .select_related("user")
        .filter(institution=institution)
        .order_by("user__first_name", "user__username")
    )

    # Edit mode: ?edit_admin=<InstitutionAdmin.id>
    edit_id = request.GET.get("edit_admin") or request.POST.get("editing_admin_id") or ""
    editing_admin = None
    if str(edit_id).isdigit():
        editing_admin = InstitutionAdmin.objects.select_related("user").filter(
            id=int(edit_id),
            institution=institution
        ).first()

    # -------------------------
    # DELETE EXISTING ADMIN (POST)
    # -------------------------
    if request.method == "POST" and request.POST.get("delete_admin") == "1":
        admin_id = (request.POST.get("admin_id") or "").strip()
        admin_obj = get_object_or_404(InstitutionAdmin, id=admin_id, institution=institution)

        u = admin_obj.user

        # safety
        if u.id == request.user.id:
            messages.error(request, "You cannot delete your own account.")
            return redirect("institution_detail", institution_id=institution.id)

        if u.is_superuser:
            messages.error(request, "You cannot delete a Super Admin account.")
            return redirect("institution_detail", institution_id=institution.id)

        username = u.username
        # delete assignment first, then user
        admin_obj.delete()
        u.delete()

        messages.success(request, f"Admin user '{username}' deleted successfully.")
        return redirect("institution_detail", institution_id=institution.id)

    # -------------------------
    # TOGGLE ACTIVE/INACTIVE (POST)
    # -------------------------
    if request.method == "POST" and "toggle_status" in request.POST:
        admin_id = request.POST.get("admin_id")
        admin = get_object_or_404(InstitutionAdmin, id=admin_id, institution=institution)
        admin.is_active = not admin.is_active
        admin.save(update_fields=["is_active"])

        messages.success(request, f"Admin user {admin.user.username} status updated.")
        return redirect("institution_detail", institution_id=institution.id)

    # -------------------------
    # CREATE or UPDATE ADMIN USER (POST)
    # -------------------------
    if request.method == "POST":
        # If editing_admin exists => update user
        if editing_admin:
            form = UserAdminForm(request.POST, instance=editing_admin.user)
            if form.is_valid():
                user = form.save(commit=False)
                user.is_staff = True
                user.is_superuser = False
                user.save()

                messages.success(request, "Admin user updated successfully.")
                return redirect("institution_detail", institution_id=institution.id)
            else:
                messages.error(request, "Please fix the errors below.")
        else:
            # Create new admin
            form = UserAdminForm(request.POST)
            if form.is_valid():
                user = form.save(commit=False)
                user.is_staff = True
                user.is_superuser = False
                user.set_password("INSADMIN2025")
                user.save()

                InstitutionAdmin.objects.create(user=user, institution=institution)

                messages.success(request, "Admin user created successfully. Default password: INSADMIN2025")
                return redirect("institution_detail", institution_id=institution.id)
            else:
                messages.error(request, "Please fix the errors below.")
    else:
        form = UserAdminForm(instance=(editing_admin.user if editing_admin else None))

    return render(
        request,
        "institution_admin/institution_detail.html",
        {
            "institution": institution,
            "form": form,
            "institution_admins": institution_admins,
            "editing_admin": editing_admin,  # used by template to show edit mode
        }
    )

    
# Edit or Create User (Admin can create or edit users)
@login_required
def create_or_edit_user(request, user_id=None):
    if not request.user.is_superuser:
        return HttpResponse("Unauthorized", status=403)

    if user_id:
        user = get_object_or_404(User, id=user_id)
        form = UserAdminForm(request.POST or None, instance=user)
    else:
        user = None
        form = UserAdminForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        is_new = (user_id is None)

        u = form.save(commit=False)
        u.is_staff = True
        u.is_superuser = False

        # set default password only on CREATE
        if is_new:
            u.set_password("INSADMIN2025")

        u.save()

        messages.success(
            request,
            "Admin user created successfully. Default password: INSADMIN2025"
            if is_new else
            "Admin user updated successfully."
        )

        # stay on this page (like your 2nd photo)
        return redirect("create_user")

    # show existing admin users on the right panel
    admin_users = User.objects.filter(is_staff=True, is_superuser=False).order_by("-date_joined")

    return render(
        request,
        "institution_admin/create_user.html",
        {"form": form, "user_id": user_id, "admin_users": admin_users},
    )


# Delete User
@login_required
def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    messages.success(request, "User deleted successfully.")
    return redirect('institution_admin_dashboard')  # Redirect to the admin dashboard after deleting a user

# Create, Edit, or Delete Institution
@login_required
def create_or_edit_institution(request, institution_id=None):
    if institution_id:
        institution = get_object_or_404(Institution, id=institution_id)
        form = InstitutionForm(request.POST or None, instance=institution)
    else:
        form = InstitutionForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        form.save()
        if institution_id:
            messages.success(request, "Institution updated successfully.")
        else:
            messages.success(request, "Institution created successfully.")
        return redirect('institution_admin_dashboard')  # Redirect to admin dashboard after creating or editing an institution

    return render(request, 'institution_admin/create_institution.html', {'form': form, 'institution_id': institution_id})  # Correct template path

# Delete Institution
@login_required
def delete_institution(request, institution_id):
    institution = get_object_or_404(Institution, id=institution_id)
    institution.delete()
    messages.success(request, "Institution deleted successfully.")
    return redirect('institution_admin_dashboard')  # Redirect to the admin dashboard after deleting an institution

# Assign Institution Admin (Admin can assign a user to an institution)
@login_required
@user_passes_test(lambda u: u.is_superuser)
def assign_institution_admin(request, admin_id=None):
    # Edit mode (if admin_id exists)
    if admin_id:
        institution_admin = get_object_or_404(InstitutionAdmin, id=admin_id)
        form = InstitutionAdminForm(request.POST or None, instance=institution_admin)
    else:
        institution_admin = None
        form = InstitutionAdminForm(request.POST or None)

    # Optional: limit dropdown to staff-admin users only (not superuser)
    if "user" in form.fields:
        form.fields["user"].queryset = User.objects.filter(
            is_staff=True, is_superuser=False
        ).order_by("username")

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(
            request,
            "Institution admin assignment updated successfully."
            if admin_id else
            "Institution admin assigned successfully."
        )
        # stay on the same page (refresh)
        return redirect("assign_institution_admin")

    # list of existing assignments
    assignments = (
        InstitutionAdmin.objects
        .select_related("user", "institution")
        .order_by("institution__name", "user__username")
    )

    context = {
        "form": form,
        "admin_id": admin_id,
        "editing": institution_admin is not None,
        "assignments": assignments,
    }
    return render(request, "institution_admin/assign_institution_admin.html", context)


# Delete Institution Admin Assignment
@login_required
@user_passes_test(lambda u: u.is_superuser)
@require_POST
def delete_institution_admin(request, admin_id):
    institution_admin = get_object_or_404(InstitutionAdmin, id=admin_id)
    institution_admin.delete()
    messages.success(request, "Institution admin assignment deleted successfully.")

    # refresh same page
    next_url = request.POST.get("next") or reverse("assign_institution_admin")
    return redirect(next_url)
def institution_admin_login(request):
    """Login for the assigned Institution Admin"""
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            try:
                institution_admin = InstitutionAdmin.objects.get(user=user)
                login(request, user)
                return redirect("institution_dashboard", institution_id=institution_admin.institution.id)
            except InstitutionAdmin.DoesNotExist:
                messages.error(request, "You are not authorized to access this page.")
                return redirect("institution_admin_login")
        else:
            messages.error(request, "Invalid username or password.")
            return redirect("institution_admin_login")

    return render(request, "institution/admin_login.html")

# Admin Logout
def institution_admin_logout(request):
    logout(request)
    return redirect('institution_admin_login')  

# @login_required
# @user_passes_test(lambda u: u.is_superuser)
# def system_backup_view(request):
#     """
#     Creates a ZIP file containing:
#     1. A JSON dump of the entire database.
#     2. All uploaded media files (scanned sheets).
#     """
#     timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
#     zip_filename = f"System_Backup_{timestamp}.zip"

#     # Create the ZIP in memory
#     buffer = BytesIO()
#     with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as backup_zip:
        
#         # --- 1. Database Backup ---
#         # We dump the database to a JSON string
#         data = StringIO()
#         try:
#             management.call_command('dumpdata', indent=2, stdout=data)
#             backup_zip.writestr('database_dump.json', data.getvalue())
#         except Exception as e:
#             messages.error(request, f"Database backup failed: {str(e)}")
#             return redirect('admin_dashboard')

#         # --- 2. Media Files Backup ---
#         media_root = Path(settings.MEDIA_ROOT)
#         if media_root.exists():
#             for file_path in media_root.rglob('*'):
#                 if file_path.is_file():
#                     # Preserve path relative to MEDIA_ROOT
#                     arcname = os.path.join('media', file_path.relative_to(media_root))
#                     backup_zip.write(file_path, arcname)

#     buffer.seek(0)
    
#     # Return the ZIP as a download
#     response = HttpResponse(buffer, content_type='application/zip')
#     response['Content-Disposition'] = f'attachment; filename={zip_filename}'
#     return response

# @login_required
# @user_passes_test(lambda u: u.is_superuser)
# def maintenance_page(request):
#     """A central hub for Backup and System Reset."""
#     institution = get_current_institution(request)
    
#     # Calculate simple stats for the UI
#     context = {
#         'institution_name': institution.name if institution else "System",
#         'total_scans': ScanResult.objects.count(),
#         'total_users': User.objects.count(),
#         'last_backup': "Never" # In a real app, you'd track this in a model
#     }
#     return render(request, "institution_admin/maintenance.html", context)

# @login_required
# @user_passes_test(lambda u: u.is_superuser)
# def reset_data_confirm(request):
#     """A dangerous view to wipe current year data only."""
#     if request.method == "POST":
#         confirm_text = request.POST.get('confirm_text')
#         if confirm_text == "ERASE":
#             # Clear operational data but keep infrastructure (Teachers/Admins)
#             ScanResult.objects.all().delete()
#             # Student.objects.all().delete() # Uncomment if you want to wipe students too
#             messages.success(request, "System data has been reset successfully.")
#             return redirect('admin_dashboard')
#         else:
#             messages.error(request, "Incorrect confirmation text.")
    
#     return render(request, "institution_admin/reset_confirm.html")

# Institution Dashboard
@login_required
def institution_dashboard(request, institution_id):
    """Dashboard for logged-in Institution Admin"""
    try:
        institution_admin = InstitutionAdmin.objects.get(user=request.user)
        institution = institution_admin.institution
    except InstitutionAdmin.DoesNotExist:
        return redirect("institution_admin_login")

    if institution.id != institution_id:
        messages.error(request, "You are not authorized to view this institution.")
        return redirect("institution_admin_login")

    academic_year = institution.school_year  # Fetch the academic year

    context = {
        "institution_name": institution.name,
        "academic_year": academic_year,
    }

    return render(request, "analysis/admin_dashboard.html", context)

# ------------------------------------------------------------------------------
# ADMIN LOGIN + DASHBOARD
# ------------------------------------------------------------------------------

def is_super_admin(user):
    return user.is_authenticated and user.is_superuser


def is_institution_admin(user):
    return user.is_authenticated and InstitutionAdmin.objects.filter(user=user).exists()


def admin_login(request):
    """
    One login page:
    - superuser -> admin_dashboard
    - assigned institution admin -> admin_dashboard (same dashboard after login)
    """
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is None:
            messages.error(request, "Invalid username or password.")
            return redirect("admin_login")

        # Superuser -> Admin Dashboard
        if user.is_superuser:
            login(request, user)
            return redirect("admin_dashboard")

        # Institution admin -> must be authorized, then Admin Dashboard
        try:
            InstitutionAdmin.objects.select_related("institution").get(user=user, is_active=True)
        except InstitutionAdmin.DoesNotExist:
            messages.error(request, "You are not authorized to access this portal.")
            return redirect("admin_login")

        login(request, user)
        return redirect("admin_dashboard")

    return render(request, "analysis/admin_login.html")
# --- optional models (avoid crash if not available) ---
try:
    from .models import ScanResult
except Exception:
    ScanResult = None

import re
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count
from django.utils import timezone
from .models import (
    Student, Teacher, ScanResult, Subject, GradingPeriod, Grade, Section
)
from .forms import GradingPeriodForm, GradeForm
from django.db.models import Prefetch

def _get_teacher_or_404(request, institution):
    return get_object_or_404(Teacher, user=request.user, institution=institution)

def _get_teacher_or_403(request, institution):
    """
    Ensures the logged-in user is a Teacher in this institution.
    Returns Teacher or JsonResponse(403).
    """
    teacher = (Teacher.objects
               .select_related("user", "institution")
               .filter(user=request.user, institution=institution)
               .first())
    if not teacher:
        return None
    return teacher

@login_required
def admin_dashboard(request):
    institution = get_current_institution(request)
    if not institution:
        messages.error(request, "No institution selected.")
        return redirect("admin_login")

    # --- SCHOOL YEAR LOGIC ---
    active_sy = getattr(institution, "school_year", "Not Set")
    viewing_year = request.GET.get("view_sy", active_sy)

    # --- SCHOOL YEAR LOGIC ---
    active_sy = getattr(institution, "school_year", "Not Set")
    viewing_year = request.GET.get("view_sy") or active_sy

    # Archives list should come from SchoolYear table (NOT Student)
    all_years = list(
        SchoolYear.objects.filter(institution=institution)
        .values_list("year", flat=True)
        .order_by("-year")
    )

    # Ensure active_sy is included even if missing in SchoolYear table
    if active_sy and active_sy != "Not Set" and active_sy not in all_years:
        all_years.insert(0, active_sy)

    school_year_rows = list(
        SchoolYear.objects.filter(institution=institution).order_by("-year")
    )

    # Ensure active_sy is reflected in list (in case institution.school_year exists but row missing)
    if active_sy and active_sy != "Not Set" and not any(sy.year == active_sy for sy in school_year_rows):
        school_year_rows.insert(0, SchoolYear(institution=institution, year=active_sy, is_active=True))

    # --- GRADING PERIODS ---
    grading_periods = (
        GradingPeriod.objects.filter(institution=institution, school_year=viewing_year)
        .order_by("start_date")
    )

    selected_gp_id = request.GET.get("gp")
    selected_grading_period = None
    if selected_gp_id:
        selected_grading_period = grading_periods.filter(id=selected_gp_id).first()

    grading_buttons = []
    for gp in grading_periods:
        grading_buttons.append(
            {
                "id": gp.id,
                "period": gp.period,
                "is_selected": str(gp.id) == str(selected_gp_id),
                "enabled": True,  # you can compute enabled later if you want
            }
        )

    # --- ARCHIVE HISTORY TREE LOGIC (GRADE -> SECTION -> SUBJECT) ---
    archive_tree = {}
    is_viewing_archive = "view_sy" in request.GET

    if is_viewing_archive:
        # 1) Prebuild FULL tree from curriculum (all grades/sections/subjects)
        all_grades = (
            Grade.objects.filter(institution=institution)
            .prefetch_related("sections", "subjects")
            .order_by("name")
        )

        for g in all_grades:
            grade_name = getattr(g, "name", None) or "Unknown Grade"
            archive_tree.setdefault(grade_name, {})

            sections = list(g.sections.all()) if hasattr(g, "sections") else []
            subjects = list(g.subjects.all()) if hasattr(g, "subjects") else []

            # show grade even if no sections
            if not sections:
                archive_tree[grade_name].setdefault("No Sections", {})
                continue

            for sec in sections:
                section_name = getattr(sec, "name", None) or "Unknown Section"
                archive_tree[grade_name].setdefault(section_name, {})

                if not subjects:
                    archive_tree[grade_name][section_name].setdefault(
                        "No Subjects",
                        {
                            "teachers": set(),
                            "records": [],
                            "grade_id": g.id,
                            "section_id": sec.id,
                            "subject_id": None,
                            "assessments": [],
                            "_assess_map": {},
                            "_student_map": {},
                            "student_rows": [],
                        },
                    )
                    continue

                for sub in subjects:
                    subject_name = getattr(sub, "name", None) or "Unknown Subject"
                    archive_tree[grade_name][section_name].setdefault(
                        subject_name,
                        {
                            "teachers": set(),
                            "records": [],
                            "grade_id": g.id,
                            "section_id": sec.id,
                            "subject_id": sub.id,
                            "assessments": [],
                            "_assess_map": {},
                            "_student_map": {},
                            "student_rows": [],
                        },
                    )

        # 2) Overlay ScanResult records for that SY (and optional GP filter)
        archive_records = (
            ScanResult.objects.filter(institution=institution, school_year=viewing_year)
            .select_related(
                "grade",
                "section",
                "subject",
                "answer_key__user",
                "answer_key__grading_period",
            )
            .order_by(
                "grade__name",
                "section__name",
                "subject__name",
                "student_name",
                "-created_at",
            )
        )

        # Apply GP filter if selected
        if selected_grading_period:
            archive_records = archive_records.filter(
                answer_key__grading_period=selected_grading_period
            )

        for record in archive_records:
            grade_name = getattr(getattr(record, "grade", None), "name", None) or "Unknown Grade"
            section_name = getattr(getattr(record, "section", None), "name", None) or "Unknown Section"
            subject_name = getattr(getattr(record, "subject", None), "name", None) or "Unknown Subject"

            ak = getattr(record, "answer_key", None)
            ak_user = getattr(ak, "user", None)

            teacher_name = (
                (ak_user.get_full_name() if ak_user else "") or
                (ak_user.username if ak_user else "Unknown Teacher")
            )

            gp_obj = getattr(ak, "grading_period", None)
            gp_label = getattr(gp_obj, "period", None) or "-"

            quiz_name = getattr(ak, "quiz_name", None) or "Assessment"

            # Ensure keys exist even if record combo not in curriculum
            archive_tree.setdefault(grade_name, {})
            archive_tree[grade_name].setdefault(section_name, {})
            archive_tree[grade_name][section_name].setdefault(
                subject_name,
                {
                    "teachers": set(),
                    "records": [],
                    "grade_id": getattr(getattr(record, "grade", None), "id", None),
                    "section_id": getattr(getattr(record, "section", None), "id", None),
                    "subject_id": getattr(getattr(record, "subject", None), "id", None),
                    "assessments": [],
                    "_assess_map": {},
                    "_student_map": {},
                    "student_rows": [],
                },
            )

            payload = archive_tree[grade_name][section_name][subject_name]

            # make sure internal structures exist
            payload.setdefault("teachers", set())
            payload.setdefault("records", [])
            payload.setdefault("assessments", [])
            payload.setdefault("_assess_map", {})
            payload.setdefault("_student_map", {})
            payload.setdefault("student_rows", [])

            if isinstance(payload["teachers"], list):
                payload["teachers"] = set(payload["teachers"])

            payload["teachers"].add(teacher_name)

            payload["records"].append(
                {
                    "student_name": record.student_name or "Unknown Student",
                    "score": record.score,
                    "max_score": record.max_score,
                    "created_at": record.created_at,
                    "grading_period": gp_label,
                    "quiz_name": quiz_name,
                }
            )

            # stable assessment key (unique per AnswerKey if possible)
            ak_id = getattr(ak, "id", None) or f"{quiz_name}|{gp_label}|{record.max_score}"
            ass_key = str(ak_id)

            if ass_key not in payload["_assess_map"]:
                payload["_assess_map"][ass_key] = {
                    "key": ass_key,
                    "quiz_name": quiz_name,
                    "grading_period": gp_label,
                    "max_score": record.max_score,
                    "created_at": record.created_at,
                }
                payload["assessments"].append(payload["_assess_map"][ass_key])

            st_name = record.student_name or "Unknown Student"
            payload["_student_map"].setdefault(st_name, {})
            payload["_student_map"][st_name][ass_key] = {
                "score": record.score,
                "max_score": record.max_score,
            }

        # 3) FINALIZE EACH PAYLOAD: teachers list + sorted assessments + build student_rows
        for gname, sections in archive_tree.items():
            for sname, subjects in sections.items():
                for subname, payload in subjects.items():
                    if not isinstance(payload, dict):
                        continue

                    # teachers set -> list
                    if isinstance(payload.get("teachers"), set):
                        payload["teachers"] = sorted(payload["teachers"])
                    elif payload.get("teachers") is None:
                        payload["teachers"] = []

                    # ensure assessments exists
                    payload.setdefault("assessments", [])
                    payload.setdefault("_student_map", {})
                    payload.setdefault("student_rows", [])

                    # sort columns
                    payload["assessments"] = sorted(
                        payload["assessments"],
                        key=lambda a: (
                            a.get("created_at") is None,
                            a.get("created_at"),
                            a.get("quiz_name") or "",
                        ),
                    )

                    # build roster (prefer Student table)
                    roster = []
                    try:
                        students_qs = Student.objects.filter(
                            institution=institution,
                            school_year=viewing_year
                        )

                        # detect FK fields safely
                        has_grade_fk = False
                        has_section_fk = False
                        try:
                            Student._meta.get_field("grade")
                            has_grade_fk = True
                        except Exception:
                            pass

                        try:
                            Student._meta.get_field("section")
                            has_section_fk = True
                        except Exception:
                            pass

                        # apply grade/section filter if those fields exist
                        if payload.get("grade_id") and has_grade_fk:
                            students_qs = students_qs.filter(grade_id=payload["grade_id"])
                        if payload.get("section_id") and has_section_fk:
                            students_qs = students_qs.filter(section_id=payload["section_id"])

                        # build names robustly (supports different Student schemas)
                        student_fields = {f.name for f in Student._meta.fields}

                        if {"first_name", "last_name"}.issubset(student_fields):
                            pairs = list(
                                students_qs.order_by("last_name", "first_name")
                                .values_list("first_name", "last_name")
                            )
                            roster = [f"{fn} {ln}".strip() for fn, ln in pairs if (fn or ln)]

                        elif "name" in student_fields:
                            roster = list(students_qs.order_by("name").values_list("name", flat=True))

                        elif "student_name" in student_fields:
                            roster = list(students_qs.order_by("student_name").values_list("student_name", flat=True))

                        else:
                            # fallback: whatever __str__ returns
                            roster = [str(s) for s in students_qs]

                    except Exception:
                        roster = []

                    # fallback: names from scan results (if Student table has none)
                    if not roster:
                        roster = sorted(payload.get("_student_map", {}).keys())

                    # build student_rows matrix
                    rows = []
                    st_map = payload.get("_student_map", {})
                    for st in roster:
                        cells = []
                        for a in payload["assessments"]:
                            key = a["key"]
                            cells.append(st_map.get(st, {}).get(key))  # None if missing
                        rows.append({"student_name": st, "cells": cells})

                    payload["student_rows"] = rows

    # --- GRADES & SECTIONS MANAGEMENT ---
    grades_qs = Grade.objects.filter(institution=institution).prefetch_related("sections", "subjects")

    def grade_number(g):
        match = re.search(r"\d+", g.name or "")
        return int(match.group()) if match else 0

    grades = sorted(grades_qs, key=grade_number)

    # --- TEACHER & ASSIGNMENT LOGIC ---
    assignment_prefetch = Prefetch(
        "class_assignments",
        queryset=TeacherClassAssignment.objects.filter(school_year=viewing_year)
        .select_related("grade", "section", "subject"),
        to_attr="current_assignments",
    )

    teachers = (
        Teacher.objects.filter(institution=institution)
        .select_related("user")
        .prefetch_related(assignment_prefetch)
        .order_by("user__last_name")
    )

    # Stats
    active_teacher_ids = (
        ScanResult.objects.filter(institution=institution, school_year=viewing_year)
        .values_list("answer_key__user_id", flat=True)
        .distinct()
    )

    total_teachers_count = teachers.count()
    active_this_year_count = teachers.filter(user_id__in=active_teacher_ids).count()

    total_students = Student.objects.filter(institution=institution, school_year=viewing_year).count()
    total_subjects = Subject.objects.filter(grade__institution=institution).count()
    total_reports = ScanResult.objects.filter(institution=institution, school_year=viewing_year).count()

    context = {
        "institution_name": getattr(institution, "name", ""),
        "teacher_name": request.user.get_full_name() or request.user.username,
        "academic_year": active_sy,
        "viewing_year": viewing_year,
        "all_years": all_years,
        "school_year_rows": school_year_rows,
        "archive_tree": archive_tree,
        "total_users": total_teachers_count,
        "active_teachers_count": active_this_year_count,
        "inactive_teachers_count": total_teachers_count - active_this_year_count,
        "total_students": total_students,
        "total_subjects": total_subjects,
        "total_reports": total_reports,
        "grading_periods": grading_periods,
        "selected_grading_period": selected_grading_period,
        "grading_buttons": grading_buttons,
        "grades": grades,
        "gp_form": GradeForm(),
        "grading_period_form": GradingPeriodForm(),
        "form": AddUserForm(institution=institution),
        "current_date": timezone.now().date(),
        "teachers": teachers,
    }

    return render(request, "analysis/admin_dashboard.html", context)

@login_required
def delete_scan_result(request, pk):
    record = get_object_or_404(ScanResult, pk=pk, institution=institution)
    record.is_active = False # Mark as inactive instead of deleting
    record.save()
    messages.success(request, "Record moved to archives.")
    return redirect('admin_dashboard')

@login_required
def update_school_year(request):
    if request.method == "POST":
        new_sy = (request.POST.get("new_sy") or "").strip()
        institution = get_current_institution(request)

        if institution and new_sy:
            # deactivate old active sy rows
            SchoolYear.objects.filter(institution=institution, is_active=True).update(is_active=False)

            # ensure this SY exists in SchoolYear table (so it shows in Archives)
            sy_obj, created = SchoolYear.objects.get_or_create(
                institution=institution,
                year=new_sy,
                defaults={
                    "is_active": True,
                    "start_date": None,  # only works if null=True, blank=True in model
                    "end_date": None,    # only works if null=True, blank=True in model
                }
            )
            if not created:
                sy_obj.is_active = True
                sy_obj.save(update_fields=["is_active"])

            # sync institution active year
            institution.school_year = new_sy
            institution.save(update_fields=["school_year"])

            messages.success(request, f"School Year updated to {new_sy}. Dashboard reset for the new term.")
        else:
            messages.error(request, "Failed to update School Year.")

    return redirect("admin_dashboard")

@login_required
def activate_school_year(request, sy_id):
    institution = get_current_institution(request)
    if not institution:
        messages.error(request, "No institution selected.")
        return redirect("admin_login")

    if request.method != "POST":
        return redirect("admin_dashboard")

    sy = get_object_or_404(SchoolYear, id=sy_id, institution=institution)

    with transaction.atomic():
        # deactivate all
        SchoolYear.objects.filter(institution=institution).update(is_active=False)

        # activate selected
        sy.is_active = True
        sy.save(update_fields=["is_active"])

        # also set institution.school_year snapshot
        institution.school_year = sy.year
        institution.save(update_fields=["school_year"])

    messages.success(request, f"Activated School Year: {sy.year}")
    return redirect("admin_dashboard")


@login_required
def deactivate_school_year(request, sy_id):
    institution = get_current_institution(request)
    if not institution:
        messages.error(request, "No institution selected.")
        return redirect("admin_login")

    if request.method != "POST":
        return redirect("admin_dashboard")

    sy = get_object_or_404(SchoolYear, id=sy_id, institution=institution)

    with transaction.atomic():
        sy.is_active = False
        sy.save(update_fields=["is_active"])

        # If you deactivated the currently active SY, clear institution.school_year
        if institution.school_year == sy.year:
            institution.school_year = "Not Set"
            institution.save(update_fields=["school_year"])

    messages.success(request, f"Deactivated School Year: {sy.year}")
    return redirect("admin_dashboard")

@login_required
def sy_history_list(request):
    """Shows a list of all School Years found in the system."""
    institution = get_current_institution(request)
    # Get unique years from Student records
    years = Student.objects.filter(institution=institution).values_list('school_year', flat=True).distinct().order_by('-school_year')
    
    return render(request, "analysis/admin_dashboard.html", {
        "years": years,
        "institution_name": institution.name
    })

@login_required
def sy_history_detail(request, year_slug):
    institution = get_current_institution(request)
    
    # Corrected '.order_by' instead of '.order_at'
    reports_qs = ScanResult.objects.filter(
        institution=institution,
        school_year__iexact=year_slug
    ).select_related(
        'answer_key__user', 
        'grade', 
        'section', 
        'subject'
    ).order_by('-created_at')  # <--- Changed this line

    history_tree = {}

    for report in reports_qs:
        # 1. Identify Teacher
        u = report.answer_key.user if (report.answer_key and report.answer_key.user) else None
        teacher_name = u.get_full_name() if (u and u.get_full_name()) else (u.username if u else "Unassigned")

        # 2. Identify Class (Grade | Subject | Section)
        g = report.grade.name if report.grade else "N/A"
        s = report.subject.name if report.subject else "N/A"
        sec = report.section.name if report.section else "N/A"
        class_label = f"{g} | {s} | Section: {sec}"

        # Initialize nesting
        if teacher_name not in history_tree:
            history_tree[teacher_name] = {}
        
        if class_label not in history_tree[teacher_name]:
            history_tree[teacher_name][class_label] = []

        # Add the student record
        history_tree[teacher_name][class_label].append(report)

    return render(request, "analysis/admin_dashboard.html", {
        "viewing_year": year_slug,
        "archive_tree": archive_tree,
        "institution_name": institution.name if institution else "System",
        "has_data": bool(history_tree)
    })
    
@login_required
def u_archives_years(request):
    institution = get_current_institution(request)
    if not institution:
        return JsonResponse({"ok": False, "error": "No institution context."}, status=400)

    teacher = _get_teacher_or_403(request, institution)
    if not teacher:
        return JsonResponse({"ok": False, "error": "Teacher not found."}, status=404)

    # Years where teacher has class assignments OR has answer keys/scans
    years_from_assign = (TeacherClassAssignment.objects
        .filter(teacher=teacher)
        .values_list("school_year", flat=True)
        .distinct()
    )

    years_from_keys = (AnswerKey.objects
        .filter(institution=institution, user=request.user)
        .values_list("school_year", flat=True)
        .distinct()
    )

    years_from_scans = (ScanResult.objects
        .filter(institution=institution, answer_key__user=request.user)
        .values_list("school_year", flat=True)
        .distinct()
    )

    years = sorted(
        set(list(years_from_assign) + list(years_from_keys) + list(years_from_scans)),
        reverse=True
    )

    return JsonResponse({"ok": True, "years": years})


@login_required
def u_archives_grades(request, sy):
    institution = get_current_institution(request)
    if not institution:
        return JsonResponse({"ok": False, "error": "No institution context."}, status=400)

    teacher = _get_teacher_or_403(request, institution)
    if not teacher:
        return JsonResponse({"ok": False, "error": "Teacher not found."}, status=404)

    grade_ids = (TeacherClassAssignment.objects
        .filter(teacher=teacher, school_year=sy)
        .values_list("grade_id", flat=True)
        .distinct()
    )

    grades = list(
        Grade.objects.filter(id__in=grade_ids)
        .values("id", "name")
        .order_by("name")
    )
    return JsonResponse({"ok": True, "sy": sy, "grades": grades})


@login_required
def u_archives_sections(request, sy, grade_id):
    institution = get_current_institution(request)
    if not institution:
        return JsonResponse({"ok": False, "error": "No institution context."}, status=400)

    teacher = _get_teacher_or_403(request, institution)
    if not teacher:
        return JsonResponse({"ok": False, "error": "Teacher not found."}, status=404)

    section_ids = (TeacherClassAssignment.objects
        .filter(teacher=teacher, school_year=sy, grade_id=grade_id)
        .values_list("section_id", flat=True)
        .distinct()
    )

    sections = list(
        Section.objects.filter(id__in=section_ids)
        .values("id", "name")
        .order_by("name")
    )
    return JsonResponse({"ok": True, "sy": sy, "grade_id": grade_id, "sections": sections})


@login_required
def u_archives_subjects(request, sy, grade_id, section_id):
    institution = get_current_institution(request)
    if not institution:
        return JsonResponse({"ok": False, "error": "No institution context."}, status=400)

    teacher = _get_teacher_or_403(request, institution)
    if not teacher:
        return JsonResponse({"ok": False, "error": "Teacher not found."}, status=404)

    subject_ids = (TeacherClassAssignment.objects
        .filter(teacher=teacher, school_year=sy, grade_id=grade_id, section_id=section_id)
        .values_list("subject_id", flat=True)
        .distinct()
    )

    subjects = list(
        Subject.objects.filter(id__in=subject_ids)
        .values("id", "name")
        .order_by("name")
    )
    return JsonResponse({
        "ok": True,
        "sy": sy,
        "grade_id": grade_id,
        "section_id": section_id,
        "subjects": subjects
    })

@login_required
def u_archives_subject_detail(request, sy, grade_id, section_id, subject_id):
    institution = get_current_institution(request)
    if not institution:
        return JsonResponse({"ok": False, "error": "No institution context."}, status=400)

    # viewer must be a teacher (since this is user portal)
    viewer = _get_teacher_or_404(request, institution)
    if not viewer:
        return JsonResponse({"ok": False, "error": "No Teacher matches the given query."}, status=404)

    # Security: viewer must be assigned to this class in that SY
    assigned_qs = TeacherClassAssignment.objects.filter(
        teacher=viewer,
        school_year=sy,
        grade_id=grade_id,
        section_id=section_id,
        subject_id=subject_id
    )
    if not assigned_qs.exists():
        return JsonResponse({"ok": False, "error": "Not assigned to this class."}, status=403)

    # Find the teacher assigned to this class (could be 1, but safe if multiple)
    # If your system allows only 1 teacher per class, you'll still get 1.
    assigned_teachers = (TeacherClassAssignment.objects
        .filter(
            school_year=sy,
            grade_id=grade_id,
            section_id=section_id,
            subject_id=subject_id
        )
        .select_related("teacher__user")
        .values(
            "teacher_id",
            "teacher__user__username",
            "teacher__user__first_name",
            "teacher__user__last_name",
        )
        .distinct()
    )

    teachers_out = []
    for t in assigned_teachers:
        full = (f"{t['teacher__user__first_name']} {t['teacher__user__last_name']}".strip())
        teachers_out.append({
            "teacher_id": t["teacher_id"],
            "name": full if full else t["teacher__user__username"],
            "username": t["teacher__user__username"],
        })

    # Labels
    grade = get_object_or_404(Grade, id=grade_id)
    section = get_object_or_404(Section, id=section_id)
    subject = get_object_or_404(Subject, id=subject_id)

    # Grading periods for that SY
    gp_qs = GradingPeriod.objects.filter(institution=institution, school_year=sy).order_by("start_date")
    grading_periods = list(gp_qs.values("id", "period", "start_date", "end_date"))
    gp_by_id = {str(gp["id"]): gp for gp in grading_periods}

    # Optional filter: specific grading period
    gp_id = request.GET.get("gp")  # can be None

    # Answer keys (assessments) for that class+SY created by the assigned teacher(s)
    # NOTE: if your AnswerKey is always created by the assigned teacher, this works.
    keys_qs = AnswerKey.objects.filter(
        institution=institution,
        school_year=sy,
        grade_id=grade_id,
        section_id=section_id,
        subject_id=subject_id,
    ).order_by("grading_period_id", "quiz_name", "id")

    if gp_id:
        keys_qs = keys_qs.filter(grading_period_id=gp_id)

    keys = list(keys_qs.values("id", "quiz_name", "grading_period_id", "user_id"))
    key_ids = [k["id"] for k in keys]

    # Students in that section + SY
    students_qs = Student.objects.filter(
        institution=institution,
        school_year=sy,
        section_id=section_id
    ).order_by("last_name", "full_name")

    students = list(students_qs.values("id", "full_name", "last_name", "middle_initial"))
    student_ids = [s["id"] for s in students]

    # Scans for those keys (and students)
    scan_qs = ScanResult.objects.filter(
        institution=institution,
        school_year=sy,
        answer_key_id__in=key_ids,
        student_id__in=student_ids,  # assumes ScanResult has student FK
        grade_id=grade_id,
        section_id=section_id,
        subject_id=subject_id,
    ).select_related("answer_key")

    # Map scans: student_id -> key_id -> scan info
    scans_map = {}
    for s in scan_qs:
        sid = str(s.student_id)
        kid = str(s.answer_key_id)
        scans_map.setdefault(sid, {})[kid] = {
            "scan_id": s.id,
            "score": s.score,
            "max_score": s.max_score,
            "created_at": s.created_at.strftime("%Y-%m-%d %H:%M"),
        }

    # Attach scans per student
    out_students = []
    for st in students:
        sid = str(st["id"])
        out_students.append({
            **st,
            "scans": scans_map.get(sid, {})  # { key_id: {...} }
        })

    # Group assessments by grading period (this is what you asked)
    # Structure:
    # grading_periods_grouped: [{gp... , assessments:[{key...}], ...}]
    grouped = {}
    for k in keys:
        gp_key = str(k["grading_period_id"] or "")
        grouped.setdefault(gp_key, []).append({
            "id": k["id"],
            "quiz_name": k["quiz_name"],
            "grading_period_id": k["grading_period_id"],
            "created_by_user_id": k["user_id"],
        })

    grading_periods_grouped = []
    # Put "No GP" group last (if any)
    for gp in grading_periods:
        gid = str(gp["id"])
        grading_periods_grouped.append({
            **gp,
            "assessments": grouped.get(gid, [])
        })

    if "" in grouped:
        grading_periods_grouped.append({
            "id": None,
            "period": "No Grading Period",
            "start_date": None,
            "end_date": None,
            "assessments": grouped[""]
        })

    return JsonResponse({
        "ok": True,
        "sy": sy,
        "grade": {"id": grade_id, "name": grade.name},
        "section": {"id": section_id, "name": section.name},
        "subject": {"id": subject_id, "name": subject.name},

        # this is the “teacher assigned” part
        "assigned_teachers": teachers_out,

        # this is the “students list + records”
        "students": out_students,

        # “records of assessments per grading period”
        "grading_periods_grouped": grading_periods_grouped,
    })
    
def _teacher_display(user):
    if not user:
        return "Unassigned"
    full = (user.get_full_name() or "").strip()
    return full if full else user.username


def _fmt_dt(dt):
    if not dt:
        return None
    dt = localtime(dt)
    return dt.strftime("%Y-%m-%d %I:%M %p")


@login_required
def archive_years(request):
    """
    Returns all school years that have archived data for THIS institution.
    Uses ScanResult as the "source of truth" (best for assessments/records).
    """
    institution = get_current_institution(request)

    years = (
        ScanResult.objects
        .filter(institution=institution)
        .exclude(Q(school_year__isnull=True) | Q(school_year__exact=""))
        .values_list("school_year", flat=True)
        .distinct()
        .order_by("-school_year")
    )

    return JsonResponse({"years": list(years)})


@login_required
def archive_grades(request, sy):
    """
    For a school year, return grades that have records in ScanResult.
    """
    institution = get_current_institution(request)

    qs = (
        ScanResult.objects
        .filter(institution=institution, school_year__iexact=sy, grade__isnull=False)
        .select_related("grade")
        .values("grade_id", "grade__name")
        .distinct()
        .order_by("grade__name")
    )

    grades = [{"id": row["grade_id"], "name": row["grade__name"]} for row in qs]
    return JsonResponse({"grades": grades})


@login_required
def archive_sections(request, sy, grade_id):
    """
    For a school year + grade, return sections that have records.
    """
    institution = get_current_institution(request)

    qs = (
        ScanResult.objects
        .filter(
            institution=institution,
            school_year__iexact=sy,
            grade_id=grade_id,
            section__isnull=False,
        )
        .select_related("section")
        .values("section_id", "section__name")
        .distinct()
        .order_by("section__name")
    )

    sections = [{"id": row["section_id"], "name": row["section__name"]} for row in qs]
    return JsonResponse({"sections": sections})


@login_required
def archive_subjects(request, sy, grade_id, section_id):
    """
    For a school year + grade + section, return subjects that have records.
    """
    institution = get_current_institution(request)

    qs = (
        ScanResult.objects
        .filter(
            institution=institution,
            school_year__iexact=sy,
            grade_id=grade_id,
            section_id=section_id,
            subject__isnull=False,
        )
        .select_related("subject")
        .values("subject_id", "subject__name")
        .distinct()
        .order_by("subject__name")
    )

    subjects = [{"id": row["subject_id"], "name": row["subject__name"]} for row in qs]
    return JsonResponse({"subjects": subjects})


@login_required
def archive_detail(request, sy, grade_id, section_id, subject_id):
    """
    Returns the teacher assigned + student list + assessment records.

    Teacher logic:
      1) If you have a formal assignment table, you can swap this logic.
      2) For now, we infer teacher from the most recent ScanResult.answer_key.user
         for that class and school year.
    """
    institution = get_current_institution(request)

    # Get all scan/assessment records for that class and SY
    reports_qs = (
        ScanResult.objects
        .filter(
            institution=institution,
            school_year__iexact=sy,
            grade_id=grade_id,
            section_id=section_id,
            subject_id=subject_id,
        )
        .select_related("answer_key__user", "grade", "section", "subject")
        .order_by("-created_at")
    )

    # Meta (grade/section/subject labels)
    first = reports_qs.first()
    meta = {
        "sy": sy,
        "grade": first.grade.name if first and first.grade else "",
        "section": first.section.name if first and first.section else "",
        "subject": first.subject.name if first and first.subject else "",
        "teacher": "Unassigned",
    }

    # Infer teacher (most recent record that has answer_key.user)
    teacher_user = None
    for r in reports_qs[:50]:  # small scan to find a teacher quickly
        if getattr(r, "answer_key", None) and getattr(r.answer_key, "user", None):
            teacher_user = r.answer_key.user
            break
    meta["teacher"] = _teacher_display(teacher_user)

    # ---- Students list ----
    # Support either Student.grade + Student.section, OR Student.section.grade pattern
    students_qs = Student.objects.filter(institution=institution, school_year__iexact=sy)

    # Try common patterns safely:
    # 1) If Student has grade_id + section_id fields
    if hasattr(Student, "grade_id") and hasattr(Student, "section_id"):
        students_qs = students_qs.filter(grade_id=grade_id, section_id=section_id)
    else:
        # 2) If Student has section FK and Section has grade FK
        if hasattr(Student, "section_id"):
            students_qs = students_qs.filter(section_id=section_id)
        # if Student has grade FK only, apply that too
        if hasattr(Student, "grade_id"):
            students_qs = students_qs.filter(grade_id=grade_id)

    # Build a display name for students
    student_list = []
    for st in students_qs.order_by("last_name", "first_name") if hasattr(Student, "last_name") else students_qs:
        if hasattr(st, "get_full_name"):
            sname = st.get_full_name()
        else:
            # fallback common fields
            fn = getattr(st, "first_name", "") or ""
            ln = getattr(st, "last_name", "") or ""
            sname = (f"{ln}, {fn}".strip(", ")).strip() or str(st)
        student_list.append({"id": st.id, "name": sname})

    # ---- Records ----
    # We’ll serialize common fields. If your ScanResult uses different names,
    # just map them here.
    records = []
    for r in reports_qs:
        # Student name: try FK then fallback
        student_name = "-"
        if hasattr(r, "student") and r.student:
            if hasattr(r.student, "get_full_name"):
                student_name = r.student.get_full_name()
            else:
                fn = getattr(r.student, "first_name", "") or ""
                ln = getattr(r.student, "last_name", "") or ""
                student_name = (f"{ln}, {fn}".strip(", ")).strip() or str(r.student)
        else:
            # sometimes ScanResult stores a raw name field
            student_name = getattr(r, "student_name", "-") or "-"

        records.append({
            "created_at": _fmt_dt(getattr(r, "created_at", None)),
            "student": student_name,
            "assessment_name": getattr(r, "assessment_name", None) or getattr(r, "test_name", None) or "Assessment",
            "score": getattr(r, "score", None),
            "max_score": getattr(r, "max_score", None) or getattr(r, "total_items", None),
        })

    return JsonResponse({
        "meta": meta,
        "students": student_list,   # <- available if you want to show the student list separately in UI
        "records": records,
    })
    
@login_required
def admin_logout(request):
    logout(request)
    return redirect("admin_login")


@login_required
# @user_passes_test(is_admin)
def admin_profile(request):
    if request.method == "POST":
        password_form = PasswordChangeForm(request.user, request.POST)
        if password_form.is_valid():
            user = password_form.save()
            update_session_auth_hash(request, user)
            messages.success(request, "Your password has been updated successfully.")
            return redirect("admin_profile")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        password_form = PasswordChangeForm(request.user)

    return render(
        request,
        "analysis/admin_profile.html",
        {"password_form": password_form},
    )

# ------------------------------------------------------------------------------
# GRADE / SECTION / SUBJECT CRUD
# ------------------------------------------------------------------------------
def get_current_institution(request):
    """
    Institution context:
    - If user is not authenticated: None
    - Teacher: use Teacher.institution (set in session)
    - Institution Admin: always their assigned institution
    - Superuser: session / URL / single fallback
    """

    # avoid: AnonymousUser id error
    if not request.user.is_authenticated:
        return None

    # TEACHER CONTEXT (most important for teacher portal)
    teacher = (
        Teacher.objects.select_related("institution")
        .filter(user=request.user)
        .first()
    )
    if teacher and teacher.institution:
        # keep session consistent for other pages
        request.session["active_institution_id"] = teacher.institution_id
        return teacher.institution

    # INSTITUTION ADMIN CONTEXT
    inst_admin = (
        InstitutionAdmin.objects.select_related("institution")
        .filter(user=request.user)
        .first()
    )
    if inst_admin:
        request.session["active_institution_id"] = inst_admin.institution_id
        return inst_admin.institution

    # SUPERUSER CONTEXT
    if request.user.is_superuser:
        iid = request.GET.get("institution_id")
        if iid:
            request.session["active_institution_id"] = int(iid)

        active_id = request.session.get("active_institution_id")
        if active_id:
            return Institution.objects.filter(id=active_id).first()

        one = Institution.objects.all()[:2]
        if len(one) == 1:
            request.session["active_institution_id"] = one[0].id
            return one[0]

    return None

@login_required
def add_grade(request):
    institution = get_current_institution(request)
    if not institution:
        return redirect("admin_dashboard")

    if request.method == "POST":
        form = GradeForm(request.POST)
        if form.is_valid():
            grade = form.save(commit=False)
            grade.institution = institution 
            grade.save()
            messages.success(request, f"Grade '{grade.name}' added successfully.")
        else:
            messages.error(request, "Error adding grade.")
    
    # Always redirect back to dashboard so stats don't disappear
    return redirect("admin_dashboard")

@login_required
def add_grading_period(request, grading_period_id=None):
    institution = get_current_institution(request)
    if not institution:
        messages.error(request, "No institution context found.")
        return redirect("admin_login")

    active_sy = getattr(institution, "school_year", None)
    
    # Check for existing instance (for updates)
    edit_id = grading_period_id or request.GET.get("edit") or request.POST.get("grading_period_id")
    editing_obj = None
    if edit_id and str(edit_id).isdigit():
        editing_obj = get_object_or_404(GradingPeriod, id=int(edit_id), institution=institution)

    if request.method == "POST":
        form = GradingPeriodForm(request.POST, instance=editing_obj)
        if form.is_valid():
            gp = form.save(commit=False)
            gp.institution = institution
            gp.school_year = active_sy
            gp.save()
            messages.success(request, f"Grading period '{gp.period}' saved successfully!")
        else:
            # ✅ Added error feedback to see WHY it's failing
            messages.error(request, f"Error saving period: {form.errors.as_text()}")
        
        # ✅ Redirect back with ?open=grading so the tab stays open in your template
        return redirect(reverse("admin_dashboard") + "?open=grading")

    return redirect(reverse("admin_dashboard") + "?open=grading")

@require_POST
@login_required
def delete_grading_period(request):
    institution = get_current_institution(request)
    if not institution:
        messages.error(request, "No institution context found.")
        return redirect("admin_dashboard")

    gp_id = (request.POST.get("grading_period_id") or "").strip()
    if not gp_id.isdigit():
        messages.error(request, "Invalid grading period.")
        return redirect("add_grading_period")

    gp = get_object_or_404(GradingPeriod, id=int(gp_id), institution=institution)
    gp.delete()

    messages.success(request, "Grading period deleted successfully!")
    return redirect("add_grading_period")

@login_required
def edit_grade(request, grade_id):
    institution = get_current_institution(request)
    grade = get_object_or_404(Grade, id=grade_id, institution=institution)
    if request.method == "POST":
        form = GradeForm(request.POST, instance=grade)
        if form.is_valid():
            form.save()
            messages.success(request, "Grade updated successfully.")
    return redirect("admin_dashboard")


@login_required
def delete_grade(request, grade_id):
    institution = get_current_institution(request)
    grade = get_object_or_404(Grade, id=grade_id, institution=institution)

    if request.method == "POST":
        grade.delete()
        messages.success(request, "Grade deleted successfully.")
        return redirect("add_grade")

    return render(request, "analysis/confirm_delete.html", {"grade": grade})


@login_required
def manage_grade(request, grade_id):
    institution = get_current_institution(request)
    grade = get_object_or_404(Grade, id=grade_id, institution=institution)
    return render(request, "analysis/manage_grade.html", {"grade": grade})

@login_required
def add_section(request, grade_id):
    institution = get_current_institution(request)
    grade = get_object_or_404(Grade, pk=grade_id, institution=institution)
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    if request.method == "POST":
        form = SectionForm(request.POST) # Ensure SectionForm exists
        if form.is_valid():
            section = form.save(commit=False)
            section.grade = grade
            section.save()

            if is_ajax:
                return JsonResponse({
                    "id": section.id,
                    "name": section.name,
                    "editUrl": f"/grades/{grade.id}/sections/{section.id}/edit/",
                    "deleteUrl": f"/grades/{grade.id}/sections/{section.id}/delete/",
                })
            return redirect("admin_dashboard")
    return HttpResponse(status=400)

@login_required
def edit_section(request, grade_id, section_id):
    institution = get_current_institution(request)
    # Security check: ensure grade belongs to institution and section belongs to grade
    section = get_object_or_404(Section, id=section_id, grade_id=grade_id, grade__institution=institution)
    
    if request.method == "POST":
        name = request.POST.get("name")
        if name:
            section.name = name
            section.save()
            return JsonResponse({"success": True, "name": section.name})
    return JsonResponse({"success": False}, status=400)

@login_required
@require_POST
def delete_section(request, grade_id, section_id):
    institution = get_current_institution(request)
    # Single query filter for security
    deleted, _ = Section.objects.filter(
        id=section_id,
        grade_id=grade_id,
        grade__institution=institution
    ).delete()

    if deleted == 0:
        return JsonResponse({"success": False, "error": "Not found"}, status=404)
    return JsonResponse({"success": True})
    
@login_required
def add_subject(request, grade_id):
    institution = get_current_institution(request)
    grade = get_object_or_404(Grade, id=grade_id, institution=institution)
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    if request.method == "POST":
        form = SubjectForm(request.POST)
        if form.is_valid():
            subject = form.save(commit=False)
            subject.grade = grade
            subject.save()

            if is_ajax:
                return JsonResponse({
                    "id": subject.id,
                    "name": subject.name,
                    "editUrl": f"/grades/{grade.id}/subjects/{subject.id}/edit/",
                    "deleteUrl": f"/grades/{grade.id}/subjects/{subject.id}/delete/",
                })
    return HttpResponse(status=400)


@login_required
def edit_subject(request, grade_id, subject_id):
    institution = get_current_institution(request)
    grade = get_object_or_404(Grade, id=grade_id, institution=institution)
    subject = get_object_or_404(Subject, id=subject_id, grade=grade)
    
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    if request.method == "POST":
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            try:
                form.save()
                if is_ajax:
                    return JsonResponse({"success": True, "name": subject.name})
                
                messages.success(request, "Subject updated successfully.")
                return redirect("add_subject", grade_id=grade.id)
            except IntegrityError:
                if is_ajax:
                    return JsonResponse({"success": False, "error": "Subject already exists."}, status=400)
                messages.error(request, f"Subject '{form.cleaned_data['name']}' already exists.")
        elif is_ajax:
            return JsonResponse({"success": False, "errors": form.errors}, status=400)

    # Standard non-AJAX fallback
    return render(request, "analysis/edit_subject.html", {"form": form, "grade": grade})

@login_required
@require_POST
def delete_subject(request, grade_id, subject_id):
    institution = get_current_institution(request)
    # Filter for security: ensure grade belongs to current institution
    subject = get_object_or_404(Subject, id=subject_id, grade_id=grade_id, grade__institution=institution)
    
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    name = subject.name
    subject.delete()

    if is_ajax:
        return JsonResponse({"success": True, "message": f"Subject {name} deleted."})

    messages.success(request, f"Subject '{name}' deleted successfully.")
    return redirect("add_subject", grade_id=grade_id)

@login_required
def student_list(request):
    institution = get_current_institution(request)
    if not institution:
        messages.error(request, "No institution context found. Please select an institution.")
        return redirect("admin_dashboard")

    # 1. Fetch scoped data
    grades = Grade.objects.filter(institution=institution).order_by("number")
    sections = Section.objects.filter(grade__institution=institution).select_related("grade").order_by("grade__number", "name")
    
    # Only students for current institution + active school year
    students = (
        Student.objects.select_related("section", "section__grade")
        .filter(institution=institution, school_year=institution.school_year)
        .order_by("section__grade__number", "section__name", "last_name", "full_name")
    )

    editing_student_id = None

    # 2. Handle POST Request (Add/Update)
    if request.method == "POST":
        student_id = request.POST.get("student_id")
        
        if student_id:
            instance = get_object_or_404(Student, pk=student_id, institution=institution)
            form = StudentForm(request.POST, instance=instance, institution=institution)
        else:
            instance = None
            form = StudentForm(request.POST, institution=institution)

        if form.is_valid():
            student = form.save(commit=False)
            
            # Double check section belongs to this institution
            if student.section.grade.institution != institution:
                messages.error(request, "Invalid section selection.")
                return redirect("student_list")

            # Duplicate Check: Case Insensitive
            dup_qs = Student.objects.filter(
                institution=institution,
                school_year=institution.school_year,
                section=student.section,
                full_name__iexact=student.full_name,
                last_name__iexact=student.last_name,
                middle_initial__iexact=student.middle_initial,
            )
            if instance:
                dup_qs = dup_qs.exclude(pk=instance.pk)

            if dup_qs.exists():
                messages.error(request, "This student is already listed in that section.")
                editing_student_id = student_id
            else:
                student.institution = institution
                student.school_year = institution.school_year
                student.grade = student.section.grade # Ensure consistency
                student.save()
                
                msg = "Student updated successfully." if instance else "Student added successfully."
                messages.success(request, msg)
                return redirect("student_list")
        else:
            messages.error(request, "Please correct the errors below.")
            editing_student_id = student_id
    else:
        form = StudentForm(institution=institution)

    # 3. Grouping Logic for the UI
    students_by_section = defaultdict(list)
    for s in students:
        students_by_section[s.section_id].append(s)

    grades_with_sections = {}
    for grade in grades:
        section_groups = []
        # Optimization: Use pre-fetched sections if possible or order here
        for section in grade.sections.all().order_by("name"):
            stu_list = students_by_section.get(section.id, [])
            # We display the section even if empty to allow users to see all sections
            section_groups.append(SimpleNamespace(grouper=section, list=stu_list))
        
        if section_groups:
            grades_with_sections[grade] = section_groups

    return render(
        request,
        "analysis/students.html",
        {
            "form": form,
            "grades": grades,
            "sections": sections,
            "grades_with_sections": grades_with_sections,
            "editing_student_id": editing_student_id,
            "institution_name": institution.name,
            "academic_year": institution.school_year,
        },
    )

@login_required
def delete_student(request, pk):
    institution = get_current_institution(request)
    student = get_object_or_404(Student, pk=pk, institution=institution)

    if request.method == "POST":
        student.delete()
        messages.success(request, "Student deleted successfully.")
    
    return redirect("student_list")

@login_required
def import_students(request):
    if request.method != "POST":
        return redirect("student_list")

    institution = get_current_institution(request) # Ensure this helper is imported/defined
    grade_id = request.POST.get("grade")
    section_id = request.POST.get("section")
    uploaded_file = request.FILES.get("student_file")

    if not institution:
        messages.error(request, "No institution context found.")
        return redirect("student_list")

    if not (grade_id and section_id and uploaded_file):
        messages.error(request, "Missing required fields for import.")
        return redirect("student_list")

    grade = get_object_or_404(Grade, pk=grade_id, institution=institution)
    section = get_object_or_404(Section, pk=section_id, grade=grade)

    filename = (uploaded_file.name or "").lower()
    created_count = 0
    updated_gender_count = 0

    # -------------------------
    # HELPERS
    # -------------------------
    def normalize_header(s: str) -> str:
        return str(s or "").strip().lower().replace(" ", "").replace("_", "")

    def clean_name_cell(raw) -> str:
        """Removes numbering (1., 1), 1-) and bullets from the start of names."""
        if raw is None: return ""
        text = str(raw).strip()
        if not text: return ""
        # Remove bullets
        text = re.sub(r"^[•\-\u2022]+\s*", "", text).strip()
        # Remove numbering like "1. ", "1) ", "1- "
        text = re.sub(r"^\s*\d+\s*[\.\)\-:]\s*", "", text).strip()
        return text

    def parse_name_to_parts(name_text: str):
        """
        Smart parsing for:
        - "Last, First MI" -> (First, Last, MI)
        - "First MI Last" -> (First, Last, MI)
        """
        name_text = (name_text or "").strip()
        if not name_text:
            return ("", "", "")

        # CASE: "Last, First MI"
        if "," in name_text:
            last_part, rest = [p.strip() for p in name_text.split(",", 1)]
            parts = rest.split()
            mi = ""
            first = ""
            if parts:
                # If last word is an initial (e.g., "A" or "A.")
                if len(parts[-1].rstrip(".")) <= 2 and len(parts) > 1:
                    mi = parts[-1].rstrip(".")
                    first = " ".join(parts[:-1])
                else:
                    first = " ".join(parts)
            return (first.strip(), last_part.strip(), mi[:1])

        # CASE: "First MI Last" or "First Last"
        parts = name_text.split()
        if len(parts) >= 3:
            # Assume middle element is MI if it's short
            if len(parts[-2].rstrip(".")) <= 2:
                return (parts[0], parts[-1], parts[-2].rstrip(".")[:1])
            return (" ".join(parts[:-1]), parts[-1], "")
        elif len(parts) == 2:
            return (parts[0], parts[1], "")
        
        return (name_text, "", "")

    # Detect gender field name in your model
    gender_field = "gender" if hasattr(Student, "gender") else ("sex" if hasattr(Student, "sex") else None)

    def upsert_student(first, last, mi, gender_value=None):
        nonlocal created_count, updated_gender_count
        if not first and not last: return

        student, created = Student.objects.get_or_create(
            institution=institution,
            school_year=institution.school_year,
            section=section,
            last_name=last,
            full_name=first, # Mapping 'first' to full_name based on your original snippet
            defaults={
                "grade": grade,
                "middle_initial": mi,
                gender_field: gender_value if gender_field else None
            },
        )

        if created:
            created_count += 1
        elif gender_field and gender_value:
            # Update gender if existing record is missing it
            current_val = getattr(student, gender_field)
            if not current_val:
                setattr(student, gender_field, gender_value)
                student.save(update_fields=[gender_field])
                updated_gender_count += 1

    # -------------------------
    # MAIN IMPORT LOGIC
    # -------------------------
    try:
        # 1. EXCEL IMPORT
        if filename.endswith((".xlsx", ".xls")):
            wb = load_workbook(uploaded_file, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows: raise ValueError("Excel file is empty.")

            header = [normalize_header(c) for c in rows[0]]
            
            def get_col(names):
                for nm in names:
                    nm_norm = normalize_header(nm)
                    if nm_norm in header: return header.index(nm_norm)
                return None

            idx_male = get_col(["male", "boys"])
            idx_female = get_col(["female", "girls"])

            # Logic: Male/Female in Separate Columns
            if idx_male is not None or idx_female is not None:
                for row in rows[1:]:
                    if idx_male is not None and row[idx_male]:
                        f, l, m = parse_name_to_parts(clean_name_cell(row[idx_male]))
                        upsert_student(f, l, m, "Male")
                    if idx_female is not None and row[idx_female]:
                        f, l, m = parse_name_to_parts(clean_name_cell(row[idx_female]))
                        upsert_student(f, l, m, "Female")
            else:
                # Logic: Traditional Columns (First, Last, etc.)
                idx_f = get_col(["firstname", "first", "name"])
                idx_l = get_col(["lastname", "last", "surname"])
                idx_mi = get_col(["mi", "middleinitial"])
                
                if idx_f is None: raise ValueError("Could not find a 'Name' or 'First Name' column.")
                
                for row in rows[1:]:
                    first = clean_name_cell(row[idx_f])
                    last = clean_name_cell(row[idx_l]) if idx_l is not None else ""
                    mi = clean_name_cell(row[idx_mi]) if idx_mi is not None else ""
                    
                    if "," in first and not last: # Handle "Last, First" in one cell
                        first, last, mi = parse_name_to_parts(first)
                    
                    upsert_student(first, last, mi)

        # 2. CSV IMPORT
        elif filename.endswith(".csv"):
            decoded = uploaded_file.read().decode("utf-8", errors="ignore").splitlines()
            reader = csv.DictReader(decoded)
            fieldnames = [normalize_header(f) for f in (reader.fieldnames or [])]
            
            has_split = "male" in fieldnames or "female" in fieldnames

            for row in reader:
                # Map row keys to normalized headers
                row_norm = {normalize_header(k): v for k, v in row.items()}
                
                if has_split:
                    m_name = clean_name_cell(row_norm.get("male") or row_norm.get("boys"))
                    f_name = clean_name_cell(row_norm.get("female") or row_norm.get("girls"))
                    if m_name:
                        f, l, mi = parse_name_to_parts(m_name)
                        upsert_student(f, l, mi, "Male")
                    if f_name:
                        f, l, mi = parse_name_to_parts(f_name)
                        upsert_student(f, l, mi, "Female")
                else:
                    first = clean_name_cell(row_norm.get("firstname") or row_norm.get("name"))
                    last = clean_name_cell(row_norm.get("lastname") or row_norm.get("surname"))
                    mi = clean_name_cell(row_norm.get("mi") or row_norm.get("middleinitial"))
                    if "," in first and not last:
                        first, last, mi = parse_name_to_parts(first)
                    upsert_student(first, last, mi)

        # 3. PDF IMPORT (Block Category Logic)
        elif filename.endswith(".pdf"):
            pdf_reader = PdfReader(uploaded_file)
            current_gender = None
            RE_NUMBERED = re.compile(r"^\s*(\d+)\s*[\.\)\-:]\s*(.*)$")

            for page in pdf_reader.pages:
                text = page.extract_text()
                if not text: continue
                
                for line in text.splitlines():
                    line = line.strip()
                    low = line.lower()

                    # Detect Category Change
                    if low in ["male", "males", "boys"]:
                        current_gender = "Male"
                        continue
                    elif low in ["female", "females", "girls"]:
                        current_gender = "Female"
                        continue

                    # If inside a gender category, look for numbered names
                    if current_gender:
                        match = RE_NUMBERED.search(line)
                        if match:
                            name_text = clean_name_cell(match.group(2))
                            if name_text:
                                f, l, mi = parse_name_to_parts(name_text)
                                upsert_student(f, l, mi, current_gender)

        else:
            messages.error(request, "Unsupported file format.")
            return redirect("student_list")

        # Success Report
        msg = f"Imported {created_count} new students."
        if updated_gender_count > 0:
            msg += f" Updated gender for {updated_gender_count} existing students."
        messages.success(request, msg)

    except Exception as e:
        messages.error(request, f"Error processing file: {str(e)}")

    return redirect("student_list")

@login_required
def user_management(request):
    institution = get_current_institution(request)
    if not institution:
        messages.error(request, "No institution context found.")
        return redirect("admin_dashboard")

    active_sy = getattr(institution, "school_year", "")

    if request.method == "POST":
        user_id = request.POST.get("user_id")

        if user_id:
            user_instance = get_object_or_404(User, id=user_id)
            form = AddUserForm(request.POST, instance=user_instance, institution=institution)
            success_msg = "User updated successfully."
        else:
            form = AddUserForm(request.POST, institution=institution)
            success_msg = "User created successfully."

        if form.is_valid():
            with transaction.atomic():
                user = form.save(institution=institution)
                teacher, _ = Teacher.objects.get_or_create(user=user, institution=institution)

                user.first_name = request.POST.get("first_name", user.first_name)
                user.last_name = request.POST.get("last_name", user.last_name)
                user.save()

                # Only assignments are school-year based
                _save_teacher_assignments(request, teacher, institution)

            messages.success(request, success_msg)
            return redirect("user_management")  # or admin_dashboard if you want
        else:
            messages.error(request, f"Error: {form.errors}")

    teachers_qs = Teacher.objects.filter(institution=institution).select_related("user").order_by("-user__date_joined")

    teachers = teachers_qs.prefetch_related(
        Prefetch(
            "class_assignments",
            queryset=TeacherClassAssignment.objects.filter(
                school_year=active_sy
            ).select_related("grade", "section", "subject"),
            to_attr="current_assignments",
        )
    )

    grades = Grade.objects.filter(institution=institution).order_by("number")

    return render(request, "analysis/admin_dashboard.html", {
        "form": AddUserForm(institution=institution),
        "teachers": teachers,
        "grades": grades,
        "institution_name": getattr(institution, "name", ""),
        "academic_year": active_sy,
    })

def _save_teacher_assignments(request, teacher, institution):
    active_sy = getattr(institution, "school_year", "")

    TeacherClassAssignment.objects.filter(
        teacher=teacher, institution=institution, school_year=active_sy
    ).delete()

    grade_ids = request.POST.getlist("assign_grade[]")
    section_ids = request.POST.getlist("assign_section[]")
    subject_ids = request.POST.getlist("assign_subject[]")

    for g, s, sub in zip(grade_ids, section_ids, subject_ids):
        if not (g and s and sub):
            continue

        # prevent subject already taken by another teacher in same grade+section+SY
        already_taken = TeacherClassAssignment.objects.filter(
            institution=institution,
            school_year=active_sy,
            grade_id=g,
            section_id=s,
            subject_id=sub
        ).exclude(teacher=teacher).exists()

        if already_taken:
            continue  # or raise a messages.error if you want

        TeacherClassAssignment.objects.create(
            teacher=teacher,
            institution=institution,
            grade_id=g,
            section_id=s,
            subject_id=sub,
            school_year=active_sy
        )

@login_required
def get_user_assignments(request, user_id):
    """API for the JS to populate the Edit Form with existing classes."""
    institution = get_current_institution(request)
    active_sy = getattr(institution, "school_year", "")
    assignments = TeacherClassAssignment.objects.filter(
        teacher__user_id=user_id, 
        school_year=active_sy
    ).values('grade_id', 'section_id', 'subject_id')
    return JsonResponse(list(assignments), safe=False)

@login_required
def get_available_subjects(request, grade_id, section_id):
    institution = get_current_institution(request)
    active_sy = getattr(institution, "school_year", "")

    teacher_id = request.GET.get("teacher_id")  # optional for edit

    qs = Subject.objects.filter(grade_id=grade_id, grade__institution=institution).order_by("name")

    taken = TeacherClassAssignment.objects.filter(
        institution=institution,
        school_year=active_sy,
        grade_id=grade_id,
        section_id=section_id,
    )
    if teacher_id:
        taken = taken.exclude(teacher_id=teacher_id)  # keep current teacher's subject available

    qs = qs.exclude(id__in=taken.values_list("subject_id", flat=True))
    return JsonResponse([{"id": s.id, "name": s.name} for s in qs], safe=False)

@login_required
def toggle_user_status(request, user_id):
    """Toggles the is_active status of a teacher."""
    user_to_toggle = get_object_or_404(User, id=user_id)
    # Check if user belongs to the same institution for security
    teacher_profile = get_object_or_404(Teacher, user=user_to_toggle, institution=get_current_institution(request))
    
    user_to_toggle.is_active = not user_to_toggle.is_active
    user_to_toggle.save()
    
    status = "activated" if user_to_toggle.is_active else "deactivated"
    messages.success(request, f"User {user_to_toggle.username} has been {status}.")
    return redirect("admin_dashboard")

@login_required
def get_sections(request, grade_id):
    institution = get_current_institution(request)
    grade = get_object_or_404(Grade, pk=grade_id, institution=institution)
    sections = Section.objects.filter(grade=grade).order_by("name")
    data = [{"id": s.id, "name": s.name} for s in sections]
    return JsonResponse(data, safe=False)

@login_required
def get_subjects(request, grade_id):
    institution = get_current_institution(request)
    grade = get_object_or_404(Grade, pk=grade_id, institution=institution)
    subjects = Subject.objects.filter(grade=grade).order_by("name")
    data = [{"id": sub.id, "name": sub.name} for sub in subjects]
    return JsonResponse(data, safe=False)

@login_required
def edit_user(request, user_id):
    institution = get_current_institution(request)
    if not institution:
        messages.error(request, "No institution context found.")
        return redirect("admin_dashboard")

    # FIX: Remove school_year from the filter so we find existing users from previous years
    teacher = get_object_or_404(
        Teacher,
        user_id=user_id,
        institution=institution
    )

    if request.method != "POST":
        return redirect("admin_dashboard")

    first_name = (request.POST.get("first_name") or "").strip()
    last_name  = (request.POST.get("last_name") or "").strip()
    username   = (request.POST.get("username") or "").strip()
    role       = (request.POST.get("role") or "teacher").strip().lower()

    if User.objects.filter(username=username).exclude(id=teacher.user_id).exists():
        messages.error(request, "Username already exists.")
        return redirect("admin_dashboard")

    u = teacher.user
    u.first_name = first_name
    u.last_name  = last_name
    u.username   = username

    if role == "admin":
        u.is_superuser = True
        u.is_staff = True
    else:
        u.is_superuser = False
        u.is_staff = False

    u.save()

    # Update Teacher to the current school year upon edit if needed
    teacher.school_year = institution.school_year
    teacher.save()

    if role == "teacher":
        _save_teacher_assignments(request, teacher, institution)
    else:
        TeacherClassAssignment.objects.filter(teacher=teacher).delete()
        teacher.grade = teacher.section = teacher.subject = None
        teacher.save()

    messages.success(request, "User updated successfully!")
    return redirect("admin_dashboard")

@login_required
def delete_user(request, user_id):
    institution = get_current_institution(request)
    if not institution:
        messages.error(request, "No institution context found.")
        return redirect("admin_dashboard")

    teacher = get_object_or_404(
        Teacher,
        user_id=user_id,
        institution=institution,
        school_year=institution.school_year,
    )

    if request.method != "POST":
        return redirect("admin_dashboard")

    if teacher.user == request.user:
        messages.error(request, "You cannot delete your own account.")
        return redirect("admin_dashboard")

    username = teacher.user.username
    teacher.user.delete()
    messages.success(request, f"User '{username}' deleted successfully.")
    return redirect("admin_dashboard")

# ------------------------------------------------------------------------------
# USER LOGIN (TEACHER SIDE)
# ------------------------------------------------------------------------------
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render

from .models import Teacher


def user_login(request):
    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        password = request.POST.get("password") or ""

        user = authenticate(request, username=username, password=password)

        if user is not None:
            if user.is_superuser:
                messages.error(request, "Admin cannot login here. Use the admin portal instead.")
                return redirect("user_login")

            # must be created/assigned by institution admin
            teacher = Teacher.objects.select_related("institution").filter(user=user).first()
            if not teacher or not teacher.institution:
                messages.error(request, "This account is not assigned to any institution. Contact your admin.")
                return redirect("user_login")

            # login
            login(request, user)

            # store institution context (optional but helpful)
            request.session["institution_id"] = teacher.institution_id
            request.session["school_year"] = teacher.institution.school_year

            # messages.success(request, f"Welcome, {user.first_name or user.username}!")
            return redirect("user_dashboard")

        messages.error(request, "Invalid username or password.")
        return redirect("user_login")

    return render(request, "analysis/users/user_login.html")


@login_required
def user_logout(request):
    logout(request)
    return redirect("user_login")



def _extract_int(value, default=9999):
    """
    Extracts the first integer found in a string. If none, returns default.
    Examples:
      "10" -> 10
      "Grade 10" -> 10
      "G9" -> 9
      None/"" -> default
    """
    if value is None:
        return default
    s = str(value)
    m = re.search(r"\d+", s)
    return int(m.group()) if m else default

def grading_order_case():
    """Helper to order grading periods logically (1st, 2nd, etc)."""
    return Case(
        When(name__icontains="1st", then=Value(1)),
        When(name__icontains="2nd", then=Value(2)),
        When(name__icontains="3rd", then=Value(3)),
        When(name__icontains="4th", then=Value(4)),
        default=Value(5),
        output_field=IntegerField(),
    )

def format_student_name_for_scan(stu: dict) -> str:
    """
    MUST match scan_document() format_student_name().
    Input is a dict because students_qs uses .values(...)
    Output: "Last, Given MI." or "Given"
    """
    given = (stu.get("full_name") or "").strip()
    mi = (stu.get("middle_initial") or "").strip()
    last = (stu.get("last_name") or "").strip()

    if last:
        base = last
        if given:
            base += f", {given}"
        if mi:
            base += f" {mi}."
        return base

    return given or "Unknown"

def get_answerkey_max(answer_key):
    """
    Returns the max score for an AnswerKey based on its saved points.
    Priority:
      1) answer_key.total_points
      2) item_points["_points"] sum
      3) answer_key.total_items
      4) len(read_answer_key_file(csv))
    """
    if not answer_key:
        return 0.0

    # 1) total_points (best)
    try:
        tp = float(getattr(answer_key, "total_points", 0) or 0)
        if tp > 0:
            return tp
    except Exception:
        pass

    # 2) item_points dict (your format: {"_answers": [...], "_points": {...}})
    raw = getattr(answer_key, "item_points", None) or {}
    if isinstance(raw, dict):
        pts = raw.get("_points")
        if isinstance(pts, dict) and pts:
            total = 0.0
            for v in pts.values():
                try:
                    total += float(v)
                except Exception:
                    total += 1.0
            if total > 0:
                return total

    # 3) total_items fallback
    try:
        ti = int(getattr(answer_key, "total_items", 0) or 0)
        if ti > 0:
            return float(ti)
    except Exception:
        pass

    # 4) last resort: read CSV
    try:
        answers = read_answer_key_file(answer_key.file)
        return float(len(answers))
    except Exception:
        return 0.0

@login_required
def user_dashboard(request):
    teacher = (
        Teacher.objects.select_related("institution")
        .filter(user=request.user)
        .first()
    )
    institution = get_current_institution(request) or getattr(teacher, "institution", None)

    # always provide these keys so template won't crash
    def _empty_ctx(msg="(No institution assigned)"):
        empty_chart = {"months": {"labels": [], "values": []}, "subjects": {"labels": [], "values": []}}
        return render(request, "analysis/users/user_dashboard.html", {
            "institution_name": msg,
            "academic_year": "",
            "chart_data_json": json.dumps(empty_chart),
            "assigned_classes_json": json.dumps([]),
            "matrix_report_json": json.dumps({}),  # NEW
            "total_reports": 0,
            "total_scans": 0,
            "total_students": 0,
            "days_left_current_grading": 0,
            "grading_buttons": [],
            "allowed_grades": [],
            "saved_keys": [],
            "selected_grading_period": None,

            # Students section dependencies
            "grades": [],
            "sections": [],
            "grades_with_sections": {},
            "form": StudentForm(),
            "editing_student_id": None,
        })

    if not teacher or not institution:
        return _empty_ctx()

    academic_year = institution.school_year
    today = timezone.localdate()

    
    # 1) Grading Periods
    
    gp_list = list(
        GradingPeriod.objects
        .filter(institution=institution, school_year=academic_year)
        .annotate(_o=grading_order_case())
        .order_by("_o", "start_date")
    )

    active_gp = next((g for g in gp_list if g.start_date <= today <= g.end_date), None)

    selected_gp_id = request.GET.get("gp")
    selected_gp = None
    if selected_gp_id and selected_gp_id.isdigit():
        selected_gp = next((g for g in gp_list if g.id == int(selected_gp_id)), None)

    if not selected_gp:
        selected_gp = active_gp or (gp_list[0] if gp_list else None)

    
    # 2) Class Assignments
    
    assigned_qs = (
        TeacherClassAssignment.objects
        .select_related("grade", "section", "subject")
        .filter(teacher=teacher, institution=institution, school_year=academic_year)
        .order_by("grade__number", "section__name", "subject__name")
    )

    # Used by Students section dropdowns
    grades = sorted(
        list({a.grade_id: a.grade for a in assigned_qs if a.grade_id}.values()),
        key=lambda g: (getattr(g, "number", 0), g.name)
    )
    sections = sorted(
        list({a.section_id: a.section for a in assigned_qs if a.section_id}.values()),
        key=lambda s: (getattr(getattr(s, "grade", None), "number", 0),
                       getattr(getattr(s, "grade", None), "name", ""), s.name)
    )

    section_ids = list(assigned_qs.values_list("section_id", flat=True).distinct())

    gender_field = "gender" if hasattr(Student, "gender") else ("sex" if hasattr(Student, "sex") else None)

    fields = ["id", "full_name", "last_name", "middle_initial", "section_id"]
    if gender_field:
        fields.append(gender_field)

    students_qs = (
        Student.objects
        .filter(section_id__in=section_ids, institution=institution)
        .values(*fields)
        .order_by("last_name", "full_name")
    )

    
    # 3) Answer Keys
    
    saved_keys = (
        AnswerKey.objects
        .filter(
            user=request.user,
            institution=institution,
            school_year=academic_year,
            grading_period=selected_gp
        )
        .select_related("grade", "section", "subject")
    )

    # group keys by composite (grade-section-subject)
    keys_by_composite = defaultdict(list)
    for k in saved_keys:
        comp = f"{k.grade_id}-{k.section_id}-{k.subject_id}"
        keys_by_composite[comp].append({
            "id": k.id,
            "name": k.quiz_name,
            "items": k.total_items,
        })

    
    # 4) Scans (map to students)
    
    scans_qs = (
        ScanResult.objects
        .filter(
            answer_key__user=request.user,
            answer_key__school_year=academic_year,
            answer_key__grading_period=selected_gp
        )
        .select_related("answer_key")
    )

    # scan_map[formatted_student_name][answer_key_id] = {"score": "x/y", "scan_id": 123}
    scan_map = {}
    for scan in scans_qs:
        name_key = (scan.student_name or "").strip() or "(No name)"
        ak_id = str(scan.answer_key_id)
        scan_map.setdefault(name_key, {})

        max_score = float(scan.max_score or 0)
        if not max_score:
            max_score = get_answerkey_max(scan.answer_key)

        scan_map[name_key][ak_id] = {
            "score": float(scan.score or 0),
            "max_score": float(max_score or 0),
            "scan_id": scan.id,
        }

    # students grouped by section (and attach scans)
    students_by_section = defaultdict(list)
    for stu in students_qs:
        sid = stu["section_id"]

        # IMPORTANT: must match scan_document() formatting exactly
        given = (stu.get("full_name") or "").strip()
        mi = (stu.get("middle_initial") or "").strip()
        last = (stu.get("last_name") or "").strip()

        if last:
            formatted_name = last
            if given:
                formatted_name += f", {given}"
            if mi:
                formatted_name += f" {mi}."
        else:
            formatted_name = given or "Unknown"

        stu["gender"] = stu.get("gender") or stu.get("sex") or ""
        
        stu["formatted_name"] = formatted_name  # NEW (useful for reports)
        stu["scans"] = scan_map.get(formatted_name, {})
        students_by_section[sid].append(stu)

    
    # 5) Build assignments json (ASSIGNMENTS used by JS)
    
    assigned_classes_data = []
    for a in assigned_qs:
        assigned_classes_data.append({
            "grade_id": a.grade_id,
            "grade_name": a.grade.name,
            "section_id": a.section_id,
            "section_name": a.section.name,
            "subject_id": a.subject_id,
            "subject_name": a.subject.name,
            "students": students_by_section.get(a.section_id, []),
        })

    
    # NEW: Matrix Reports JSON (for your Reports icon modal)
    # matrix_report_json[compositeKey] = { meta, assessments[], students[] }
    
    matrix_report = {}

    for a in assigned_qs:
        comp = f"{a.grade_id}-{a.section_id}-{a.subject_id}"
        assessments = keys_by_composite.get(comp, [])

        # only students from THIS section
        stus = students_by_section.get(a.section_id, [])

        # build student rows with scores per assessment id
        rows = []
        for s in stus:
            row_scores = {}
            for q in assessments:
                qid = str(q["id"])
                scan_obj = (s.get("scans") or {}).get(qid)
                row_scores[qid] = scan_obj["score"] if scan_obj else ""  # empty if not scanned

            rows.append({
                "id": s["id"],
                "name": s.get("formatted_name") or (s.get("full_name") or "Unknown"),
                "gender": s.get("gender") or "",
                "scores": row_scores,
            })
        matrix_report[comp] = {
            "grade_id": a.grade_id,
            "section_id": a.section_id,
            "subject_id": a.subject_id,
            "grade_name": a.grade.name,
            "section_name": a.section.name,
            "subject_name": a.subject.name,
            "assessments": assessments,
            "students": rows,
        }

    
    # 6) Students modal content: grades_with_sections
    
    grades_with_sections = {}
    for g in grades:
        g_sections = [s for s in sections if getattr(s, "grade_id", None) == g.id]
        section_groups = []
        for sec in g_sections:
            section_groups.append({
                "grouper": sec,
                "list": students_by_section.get(sec.id, []),
            })
        grades_with_sections[g] = section_groups

    
    # 7) Stats & chart
    
    all_time_scans = ScanResult.objects.filter(
        answer_key__user=request.user,
        answer_key__school_year=academic_year
    )

    activity_data = (
        all_time_scans.filter(created_at__year=today.year)
        .values("created_at__month")
        .annotate(count=Count("id"))
        .order_by("created_at__month")
    )

    subject_data = (
        all_time_scans.values("answer_key__subject__name")
        .annotate(count=Count("id"))
    )

    chart_data = {
        "months": {
            "labels": [timezone.datetime(2000, x["created_at__month"], 1).strftime("%b") for x in activity_data],
            "values": [x["count"] for x in activity_data],
        },
        "subjects": {
            "labels": [x["answer_key__subject__name"] for x in subject_data],
            "values": [x["count"] for x in subject_data],
        },
    }

    grading_buttons = [
        {
            "id": gp.id,
            "period": getattr(gp, "period", getattr(gp, "name", "Unknown")),
            "enabled": (today >= gp.start_date),
            "is_selected": bool(selected_gp and gp.id == selected_gp.id),
        }
        for gp in gp_list
    ]

    days_left = 0
    if selected_gp and selected_gp.end_date:
        days_left = max(0, (selected_gp.end_date - today).days)

    return render(request, "analysis/users/user_dashboard.html", {
        "institution_name": institution.name,
        "academic_year": academic_year,

        "total_reports": all_time_scans.values("answer_key_id").distinct().count(),
        "total_scans": all_time_scans.count(),
        "total_students": students_qs.count(),

        "days_left_current_grading": days_left,
        "grading_buttons": grading_buttons,

        "assigned_classes_json": json.dumps(assigned_classes_data, default=str),

        # NEW (Reports modal data)
        "matrix_report_json": json.dumps(matrix_report, default=str),

        # Answer key modal needs this
        "allowed_grades": grades,
        "saved_keys": saved_keys,
        "selected_grading_period": selected_gp,

        "chart_data_json": json.dumps(chart_data),

        # Students manage records section needs these
        "grades": grades,
        "sections": sections,
        "grades_with_sections": grades_with_sections,
        "form": StudentForm(),
        "editing_student_id": None,
    })
    
def _to_list(x):
    """
    Accepts list OR string like 'ABCD' OR 'A,B,C' OR JSON-like already.
    Returns list of answers.
    """
    if x is None:
        return []
    if isinstance(x, list):
        return x
    if isinstance(x, str):
        s = x.strip()
        if not s:
            return []
        # if comma separated
        if "," in s:
            return [a.strip() for a in s.split(",") if a.strip()]
        # treat as continuous letters (ABCD...)
        return list(s)
    return []

@login_required
def user_tally_per_item(request, answer_key_id):
    # security: only allow tallies for this teacher's answer keys
    ak = get_object_or_404(AnswerKey, pk=answer_key_id, user=request.user)

    scans = (
        ScanResult.objects
        .filter(answer_key=ak)
        .order_by("id")
    )

    total_items = int(getattr(ak, "total_items", 0) or 0)

    # Try to read correct answers from AnswerKey (adjust if your field differs)
    correct_raw = (
        getattr(ak, "correct_answers", None) or
        getattr(ak, "answer_key", None) or
        getattr(ak, "answers", None)
    )
    correct = _to_list(correct_raw)

    # If ak.total_items exists but correct list is shorter, pad to total_items
    if total_items and len(correct) < total_items:
        correct = correct + [""] * (total_items - len(correct))

    # If total_items not set, infer from correct answers
    if not total_items:
        total_items = len(correct)

    correct_counts = [0] * total_items
    scanned_count = 0

    for s in scans:
        # Prefer item correctness list if you already store it
        item_correctness = (
            getattr(s, "item_correctness", None) or
            getattr(s, "item_results", None) or
            getattr(s, "correctness", None)
        )

        if item_correctness is not None:
            ic = _to_list(item_correctness)
            if len(ic) < total_items:
                ic = ic + [""] * (total_items - len(ic))
            # normalize truthy
            for i in range(total_items):
                v = ic[i]
                is_true = (v is True) or (str(v).strip().lower() in ("1", "true", "t", "yes", "y"))
                if is_true:
                    correct_counts[i] += 1
            scanned_count += 1
            continue

        # Otherwise compare student answers vs correct answers
        student_raw = (
            getattr(s, "student_answers", None) or
            getattr(s, "answers", None) or
            getattr(s, "responses", None)
        )
        student = _to_list(student_raw)
        if total_items and len(student) < total_items:
            student = student + [""] * (total_items - len(student))

        # if we can't read answers, skip
        if not student:
            continue

        scanned_count += 1
        for i in range(total_items):
            if i < len(student) and i < len(correct) and str(student[i]).strip() and str(correct[i]).strip():
                if str(student[i]).strip().upper() == str(correct[i]).strip().upper():
                    correct_counts[i] += 1

    items = []
    for i in range(total_items):
        total = scanned_count
        c = correct_counts[i]
        pct = round((c / total) * 100, 2) if total else 0
        items.append({
            "item": i + 1,
            "correct": c,
            "total": total,
            "percent": pct
        })

    return JsonResponse({
        "answer_key_id": ak.id,
        "quiz_name": getattr(ak, "quiz_name", ""),
        "total_items": total_items,
        "scanned_count": scanned_count,
        "items": items
    })
    
@require_POST
@login_required
def upload_answer_key(request):
    teacher = Teacher.objects.select_related("institution").filter(user=request.user).first()
    institution = get_current_institution(request) or getattr(teacher, "institution", None)

    quiz_name   = request.POST.get("quiz_name")
    gp_id       = request.POST.get("grading_period")
    grade_id    = request.POST.get("grade")
    section_id  = request.POST.get("section")
    subject_id  = request.POST.get("subject")
    uploaded_file = request.FILES.get("answer_key")

    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    if not all([quiz_name, gp_id, grade_id, section_id, subject_id]):
        msg = "Missing required fields for upload."
        if is_ajax:
            return JsonResponse({"ok": False, "error": msg}, status=400)
        messages.error(request, msg)
        return redirect("user_dashboard")

    try:
        answers_raw = read_answer_key_file(uploaded_file) if uploaded_file else []

        # find existing assessment with SAME quiz_name
        existing_key = AnswerKey.objects.filter(
            user=request.user,
            institution=institution,
            school_year=institution.school_year if institution else "",
            grading_period_id=gp_id,
            grade_id=grade_id,
            section_id=section_id,
            subject_id=subject_id,
            quiz_name=quiz_name,
        ).first()

        if not existing_key and not uploaded_file:
            raise ValueError("A CSV file is required for new answer keys.")

        item_count = len(answers_raw) if uploaded_file else int(existing_key.total_items or 0)

        defaults = {
            "total_points": float(item_count) if uploaded_file else float(existing_key.total_points or 0.0),
        }

        if uploaded_file:
            defaults["file"] = uploaded_file
            defaults["item_points"] = {
                "_answers": [str(a).strip().upper() for a in answers_raw],
                "_points": {str(i + 1): 1.0 for i in range(item_count)},
            }
            defaults["total_items"] = item_count

        answer_key, created = AnswerKey.objects.update_or_create(
            user=request.user,
            institution=institution,
            school_year=institution.school_year if institution else "",
            grading_period_id=gp_id,
            grade_id=grade_id,
            section_id=section_id,
            subject_id=subject_id,
            quiz_name=quiz_name,
            defaults=defaults,
        )

        msg = "Answer Key saved!" if created else "Answer Key updated!"
        messages.success(request, f"{msg} Assessment: {quiz_name}")

        # URLs for next actions
        download_url = reverse("download_bubble_sheets_for_answer_key", args=[answer_key.pk])
        review_url = reverse("answer_key_review", args=[answer_key.pk])

        if is_ajax:
            return JsonResponse({
                "ok": True,
                "message": f"{msg} ({quiz_name})",
                "answer_key_id": answer_key.pk,
                "total_items": int(answer_key.total_items or 0),
                "download_url": download_url,
                "review_url": review_url,
            })

        # fallback (non-AJAX) behavior
        return redirect("answer_key_review", pk=answer_key.pk)

    except Exception as e:
        if is_ajax:
            return JsonResponse({"ok": False, "error": str(e)}, status=400)
        messages.error(request, f"Error: {str(e)}")
        return redirect("user_dashboard")

def _pick_template_pdf(total_items: int) -> Path:

    pdf_dir = Path(getattr(settings, "PDF_TEMPLATES_DIR", ""))

    if not pdf_dir.exists():
        raise Http404(f"PDF templates folder not found: {pdf_dir}")

    # List all template PDFs like 1-10items.pdf, 1-15items.pdf, ...
    files = sorted(pdf_dir.glob("1-*items.pdf"))

    if not files:
        raise Http404(f"No templates found in: {pdf_dir}")

    # Extract N from filenames
    available = []
    for f in files:
        name = f.name  # e.g. "1-35items.pdf"
        try:
            n = int(name.split("-")[1].replace("items.pdf", ""))
            available.append((n, f))
        except Exception:
            continue

    if not available:
        raise Http404("No valid template filenames found. Use format: 1-35items.pdf")

    available.sort(key=lambda x: x[0])

    # 1) Exact match
    for n, f in available:
        if n == total_items:
            return f

    # 2) Next higher template
    for n, f in available:
        if n >= total_items:
            return f

    # 3) Fallback to highest template
    return available[-1][1]


@login_required
def download_bubble_sheets_for_answer_key(request, pk):
    answer_key = get_object_or_404(AnswerKey, pk=pk, user=request.user)

    total_items = int(answer_key.total_items or 0)
    if total_items <= 0:
        raise Http404("Answer key has no items. Upload a CSV first.")

    template_path = _pick_template_pdf(total_items)

    # THIS is where pdf_bytes comes from when using templates:
    with open(template_path, "rb") as f:
        pdf_bytes = f.read()

    filename = f"BubbleSheet_{answer_key.grade.name}_{answer_key.section.name}_{answer_key.quiz_name}_{total_items}items.pdf"
    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
    
@login_required
def get_sections(request, grade_id):
    institution = get_current_institution(request)
    if not institution:
        return JsonResponse([], safe=False)

    # grade must be from this institution
    grade = get_object_or_404(Grade, pk=grade_id, institution=institution)

    sections = Section.objects.filter(grade=grade).order_by("name")
    data = [{"id": s.id, "name": s.name} for s in sections]
    return JsonResponse(data, safe=False)


@login_required
def get_subjects(request, grade_id):
    institution = get_current_institution(request)
    if not institution:
        return JsonResponse([], safe=False)

    # grade must be from this institution
    grade = get_object_or_404(Grade, pk=grade_id, institution=institution)

    subjects = Subject.objects.filter(grade=grade).order_by("name")
    data = [{"id": s.id, "name": s.name} for s in subjects]
    return JsonResponse(data, safe=False)

@login_required
def answer_key_review(request, pk):
    institution = get_current_institution(request)
    answer_key = get_object_or_404(AnswerKey, pk=pk, user=request.user)

    options = ["A", "B", "C", "D", "E"]
    raw_data = answer_key.item_points or {}
    
    # Standardize data access
    current_answers = raw_data.get("_answers", [])
    current_points = raw_data.get("_points", {})

    if request.method == "POST":
        new_answers = []
        new_points = {}
        
        # Loop based on the number of items currently in the key
        for i in range(1, len(current_answers) + 1):
            ans = request.POST.get(f"ans_{i}", "").strip().upper()
            new_answers.append(ans if ans in options else "")
            
            pts_val = request.POST.get(f"points_{i}", "1")
            try:
                new_points[str(i)] = float(pts_val)
            except ValueError:
                new_points[str(i)] = 1.0

        answer_key.item_points = {"_answers": new_answers, "_points": new_points}
        answer_key.total_points = sum(new_points.values())
        answer_key.save()

        # Trigger background processing if items changed
        _rescore_all_scans_for_answer_key(answer_key)
        _reannotate_all_scans_for_answer_key(answer_key)

        messages.success(request, "Answer key updated.")
        return redirect("answer_key_review", pk=pk)

    # Prepare items for template display
    items = []
    for i, ans in enumerate(current_answers, 1):
        items.append({
            "number": i,
            "correct": ans,
            "points": current_points.get(str(i), 1.0)
        })
    section_students = Student.objects.filter(section=answer_key.section).order_by('last_name')

    return render(request, "analysis/users/answer_key_review.html", {
        "answer_key": answer_key,
        "items": items,
        "split_point": (len(items) + 1) // 2,
        "options": options,
        "institution_name": institution.name if institution else "",
        'section_students': section_students,
    })

@login_required
def view_answer_key(request, pk):
    institution = get_current_institution(request)
    if not institution:
        messages.error(request, "No institution context found.")
        return redirect("user_dashboard")

    answer_key = get_object_or_404(
        AnswerKey,
        pk=pk,
        user=request.user,
        institution=institution,
        school_year=institution.school_year,
    )

    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if is_ajax:
        return render(
            request,
            "analysis/users/partials/answer_key_view_inner.html",
            {"answer_key": answer_key},
        )

    return redirect("user_dashboard")

@login_required
def convert_all_answer_keys(request):
    answer_keys = AnswerKey.objects.all()

    pdf_dir = os.path.join(settings.MEDIA_ROOT, "bubble_sheets")
    os.makedirs(pdf_dir, exist_ok=True)

    pdf_files = []

    for ak in answer_keys:
        if ak.file:
            answers = read_word_answer_key(ak.file.path)
            pdf_path = os.path.join(pdf_dir, f"bubble_sheet_{ak.id}.pdf")
            generate_bubble_sheet_pdf(answers, pdf_path)
            pdf_files.append(pdf_path)

    import zipfile

    zip_path = os.path.join(pdf_dir, "all_bubble_sheets.zip")
    with zipfile.ZipFile(zip_path, "w") as zipf:
        for pdf_file in pdf_files:
            zipf.write(pdf_file, os.path.basename(pdf_file))

    return FileResponse(
        open(zip_path, "rb"),
        as_attachment=True,
        filename="all_bubble_sheets.zip",
    )

@login_required
def user_profile(request):
    if request.method == "POST":
        password_form = PasswordChangeForm(request.user, request.POST)
        if password_form.is_valid():
            user = password_form.save()
            update_session_auth_hash(request, user)
            messages.success(request, "Your password has been updated successfully.")
            return redirect("user_profile")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        password_form = PasswordChangeForm(request.user)

    return render(
        request,
        "analysis/users/user_profile.html",
        {"password_form": password_form},
    )
    
@login_required
def scan_result_modal(request, scan_id):
    scan = get_object_or_404(
        ScanResult.objects.select_related("answer_key", "grade", "section", "subject"),
        pk=scan_id,
        answer_key__user=request.user
    )

    answer_key = scan.answer_key
    answer_map, points_map = _build_omr_answer_key(answer_key)

    # rebuild items (similar logic to your scan_result page)
    items = []
    if answer_map:
        for q_no in sorted(answer_map.keys()):
            correct = answer_map.get(q_no, "")
            student_ans = ""
            status = "blank"
            is_correct = False

            if scan.answers and (q_no - 1) < len(scan.answers):
                student_ans = scan.answers[q_no - 1] or ""

            if student_ans and student_ans != "INVALID":
                if str(student_ans).strip().upper() == str(correct).strip().upper():
                    status = "correct"
                    is_correct = True
                else:
                    status = "wrong"
            else:
                status = "invalid" if student_ans == "INVALID" else "blank"

            items.append({
                "number": q_no,
                "correct": correct,
                "student": "" if student_ans == "INVALID" else student_ans,
                "is_correct": is_correct,
                "status": status,
                "points": points_map.get(q_no, 1.0),
            })

    # image_url
    image_url = ""
    if getattr(scan, "sheet_image", None):
        try:
            image_url = scan.sheet_image.url  # ImageField
        except Exception:
            # string path
            rel = str(scan.sheet_image).replace("\\", "/")
            image_url = settings.MEDIA_URL + rel

    score = int(scan.score) if scan.score is not None else 0
    max_score = int(scan.max_score) if scan.max_score is not None else 0
    score_percent = (score / max_score * 100.0) if max_score else 0.0

    return render(request, "analysis/users/partials/scan_result_modal_content.html", {
        "scan": scan,
        "answer_key": answer_key,
        "grade": scan.grade,
        "section": scan.section,
        "subject": scan.subject,
        "student_name": scan.student_name,
        "items": items,
        "score": score,
        "max_score": max_score,
        "score_percent": score_percent,
        "image_url": image_url,
    })

from itertools import groupby

@login_required
def user_capture(request, grade_id, section_id, subject_id):
    institution = get_current_institution(request)
    if not institution:
        messages.error(request, "No institution context found.")
        return redirect("user_dashboard")

    today = timezone.localdate()
    academic_year = getattr(institution, "school_year", "")

    grade = get_object_or_404(Grade, pk=grade_id)
    section = get_object_or_404(Section, pk=section_id)
    subject = get_object_or_404(Subject, pk=subject_id)

    # 1. Fetch Grading Periods
    gp_list = list(
        GradingPeriod.objects
        .filter(institution=institution, school_year=academic_year)
        .annotate(_o=grading_order_case())
        .order_by("_o", "start_date")
    )

    selected_gp = None
    active_gp = next((g for g in gp_list if g.start_date and g.end_date and g.start_date <= today <= g.end_date), None)

    gp_id = (request.GET.get("gp") or "").strip()
    if gp_id.isdigit():
        selected_gp = next((g for g in gp_list if g.id == int(gp_id)), None)

    if not selected_gp:
        selected_gp = active_gp or (gp_list[0] if gp_list else None)

    # 2. Build Sidebar/Tabs for Grading Periods
    grading_periods = []
    for gp in gp_list:
        grading_periods.append({
            "id": gp.id,
            "label": gp.period,
            "enabled": not (gp.start_date and gp.start_date > today),
            "is_active": (active_gp and gp.id == active_gp.id),
            "is_selected": (selected_gp and gp.id == selected_gp.id),
        })

    # 3. Fetch ALL AnswerKeys and Grouped Scan Results
    grouped_scans = []
    all_answer_keys = AnswerKey.objects.filter(
        user=request.user,
        institution=institution,
        school_year=academic_year,
        grading_period=selected_gp,
        grade_id=grade_id,
        section_id=section_id,
        subject_id=subject_id,
    ).order_by("-uploaded_at")
    
    selected_quiz_id = request.GET.get("quiz")
    active_quiz = None
    if selected_quiz_id and selected_quiz_id.isdigit():
        active_quiz = all_answer_keys.filter(id=int(selected_quiz_id)).first()
    
    # Default to the most recent quiz if none selected
    if not active_quiz:
        active_quiz = all_answer_keys.first()

    # 5. Filter grouped_scans to ONLY show the active quiz
    grouped_scans = []
    if active_quiz:
        scans = ScanResult.objects.filter(answer_key=active_quiz).order_by("-created_at")
        grouped_scans.append({
            "answer_key": active_quiz,
            "scans": list(scans)
        })
        
    scans_qs = ScanResult.objects.filter(
        answer_key__in=all_answer_keys
    ).select_related('answer_key').order_by("-answer_key__uploaded_at", "-created_at")

    # Group scans by AnswerKey in Python
    # This allows the template to loop through each Quiz record
    for ak in all_answer_keys:
        grouped_scans.append({
            "answer_key": ak,
            "scans": [s for s in scans_qs if s.answer_key_id == ak.id]
        })

    # For the "Scan New Sheet" button in the header, we'll pick the most recent AK
    primary_answer_key = all_answer_keys.first()

    return render(request, "analysis/users/user_capture.html", {
        "grade": grade,
        "section": section,
        "subject": subject,
        "all_answer_keys": all_answer_keys,
        "active_quiz": active_quiz,
        "grouped_scans": grouped_scans,
        "primary_answer_key": active_quiz,
        "grading_periods": grading_periods,
        "selected_grading_period": selected_gp,
        "active_grading_period": active_gp,
        "gp_id": selected_gp.id if selected_gp else "",
        "academic_year": academic_year,
    })


@login_required
def user_report(request):
    institution = get_current_institution(request)
    if not institution:
        return redirect("user_dashboard")
        
    today = timezone.localdate()
    
    # 1. Get the list of all years the teacher has data for to show in a dropdown
    available_years = AnswerKey.objects.filter(
        user=request.user, 
        institution=institution
    ).values_list('school_year', flat=True).distinct().order_by('-school_year')

    # 2. Determine which year to view: defaults to institution's active year
    active_sy = getattr(institution, "school_year", "")
    view_year = request.GET.get("sy") or active_sy

    # 3. Filter Grading periods by the selected viewing year
    grading_periods = list(
        GradingPeriod.objects
        .filter(institution=institution, school_year=view_year)
        .annotate(_ord=grading_order_case())
        .order_by("_ord", "start_date", "id")
    )

    selected_gp = None
    gp_id = (request.GET.get("gp") or "").strip()
    if gp_id.isdigit():
        selected_gp = next((gp for gp in grading_periods if gp.id == int(gp_id)), None)

    if not selected_gp and grading_periods:
        # Fallback to current date if in same year, else first period
        if view_year == active_sy:
            selected_gp = get_grading_period_for_date(institution, today)
        if not selected_gp:
            selected_gp = grading_periods[0]

    # 4. Filter AnswerKeys by the SELECTED viewing year
    qs = (
        AnswerKey.objects
        .filter(
            user=request.user, 
            institution=institution,
            school_year=view_year 
        )
        .select_related("grade", "section", "subject", "grading_period")
    )

    if selected_gp:
        qs = qs.filter(grading_period=selected_gp)

    # Search logic
    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(
            Q(grade__name__icontains=q) |
            Q(section__name__icontains=q) |
            Q(subject__name__icontains=q)
        )

    report_rows = []
    for ak in qs:
        scans_qs = ScanResult.objects.filter(answer_key=ak)
        total_scans_for_gp = scans_qs.count()
        if total_scans_for_gp == 0:
            continue 

        total_students = scans_qs.values_list("student_name", flat=True).distinct().count()
        last_scan = scans_qs.order_by("-created_at").values_list("created_at", flat=True).first()

        answer_map, points_map = _build_omr_answer_key(ak)
        total_items = max(answer_map.keys()) if answer_map else 0
        total_points = sum(float(v) for v in points_map.values()) if points_map else 0.0

        report_rows.append({
            "ak": ak,
            "total_students": total_students,
            "total_items": total_items,
            "total_points": total_points,
            "last_scan": last_scan,
            "total_scans": total_scans_for_gp,
        })

    return render(request, "analysis/users/user_report.html", {
        "report_rows": report_rows,
        "total_reports": len(report_rows),
        "total_scans": sum(r["total_scans"] for r in report_rows),
        "institution_name": institution.name,
        "academic_year": active_sy,
        "viewing_year": view_year,
        "available_years": available_years,
        "grading_buttons": build_gp_buttons(grading_periods, selected_gp),
        "selected_gp": selected_gp,
        "q": q,
    })

@login_required
def edit_answer_key(request, pk):
    # 1. Fetch the specific record
    key = get_object_or_404(AnswerKey, pk=pk, user=request.user)

    if request.method != "POST":
        return redirect("user_dashboard")

    # 2. Capture all form data
    quiz_name = request.POST.get("quiz_name")
    grade_id = request.POST.get("grade")
    section_id = request.POST.get("section")
    subject_id = request.POST.get("subject")
    uploaded_file = request.FILES.get("answer_key")  # The CSV file

    # 3. Validation
    if not all([quiz_name, grade_id, section_id, subject_id]):
        messages.error(request, "All fields are required.")
        return redirect("user_dashboard")

    try:
        # 4. Update basic metadata
        key.quiz_name = quiz_name
        key.grade_id = grade_id
        key.section_id = section_id
        key.subject_id = subject_id

        # 5. Handle File Update (The Answer Key data)
        if uploaded_file:
            # Parse the new CSV
            answers_raw = read_answer_key_file(uploaded_file)
            item_count = len(answers_raw)
            
            # Update the file field and the JSON data
            key.file = uploaded_file
            key.total_items = item_count
            key.total_points = float(item_count)
            key.item_points = {
                "_answers": [str(a).strip().upper() for a in answers_raw],
                "_points": {str(i+1): 1.0 for i in range(item_count)}
            }

        # 6. Final Save
        key.save()
        messages.success(request, f"Assessment '{quiz_name}' updated successfully.")
        
    except Exception as e:
        messages.error(request, f"Error updating answer key: {str(e)}")

    return redirect("user_dashboard")

@login_required
def delete_answer_key(request, pk):
    key = get_object_or_404(AnswerKey, pk=pk, user=request.user)
    if request.method == "POST":
        key.delete()
        messages.success(request, "Upload deleted successfully.")
    return redirect("upload_answer_key")


@login_required
def list_saved_answer_keys(request):
    grades = Grade.objects.all().order_by("number")

    grouped = {}  # { Grade: { Section: [AnswerKey, ...] } }

    for grade in grades:
        sections = Section.objects.filter(grade=grade).order_by("name")
        grouped[grade] = {}

        for section in sections:
            keys = AnswerKey.objects.filter(
                user=request.user, grade=grade, section=section
            ).select_related("subject")
            if keys.exists():
                grouped[grade][section] = keys

    return render(
        request,
        "analysis/users/saved_answer_keys.html",
        {"grouped": grouped},
    )


# ------------------------------------------------------------------------------
# WIA SCANNER JSON API (OPTIONAL, FOR FLATBED SCANNER)
# ------------------------------------------------------------------------------


# ------------------------------------------------------------------------------
# KNN OMR HELPERS (ANSWER MAP + FOLDER NAMES)
# ------------------------------------------------------------------------------

def _safe_folder_name(name: str) -> str:
    """
    Convert strings like 'Grade 7 - Lotus' into safe folder names.
    """
    if not name:
        return "unknown"
    name = name.strip()
    name = re.sub(r"[^A-Za-z0-9]+", "_", name)
    name = name.strip("_")
    return name or "unknown"


def _build_omr_answer_key(answer_key, max_items: int = 120):
    given = (getattr(stu, "full_name", "") or "").strip()
    mi = (getattr(stu, "middle_initial", "") or "").strip()
    last = (getattr(stu, "last_name", "") or "").strip()
    if last:
        base = last
        if given:
            base += f", {given}"
        if mi:
            base += f" {mi}."
        return base
    return given or "Unknown"

def _build_omr_answer_key(answer_key, max_items: int = 100):
    """Updated to 100 items to prevent the 50-item cutoff."""
    raw = getattr(answer_key, "item_points", None) or {}
    answers_list = raw.get("_answers") if isinstance(raw, dict) else []
    points_raw = raw.get("_points") if isinstance(raw, dict) else raw

    valid_opts = {"A", "B", "C", "D", "E"}
    answer_map = {}
    for idx, letter in enumerate(answers_list, start=1):
        if idx > max_items: break
        letter = (letter or "").strip().upper()
        if letter in valid_opts:
            answer_map[idx] = letter

    points_map = {}
    for k, v in (points_raw or {}).items():
        try:
            points_map[int(k)] = float(v)
        except (ValueError, TypeError): continue
    
    for i in answer_map.keys():
        points_map.setdefault(i, 1.0)
    return answer_map, points_map


# ------------------------------------------------------------------------------
# MAIN BUBBLE SHEET SCAN VIEW (SAVES ScanResult)
# ------------------------------------------------------------------------------
def get_grading_period_for_date(institution, d):
    """
    Returns the grading period where date d belongs (start_date <= d <= end_date),
    ordered 1st->4th.
    """
    if not institution or not d:
        return None

    if hasattr(d, "date"):
        d = d.date()

    order_case = Case(
        When(period="1st Grading", then=1),
        When(period="2nd Grading", then=2),
        When(period="3rd Grading", then=3),
        When(period="4th Grading", then=4),
        default=99,
        output_field=IntegerField(),
    )

    return (
        GradingPeriod.objects
        .filter(institution=institution, start_date__lte=d, end_date__gte=d)
        .annotate(_o=order_case)
        .order_by("_o")
        .first()
    )   
    
@login_required
def scan_document(request, pk):
    """
    Scan a bubble sheet for AnswerKey (pk).
    Duplicate rule: same student can be scanned again if DIFFERENT grading period.
    ❌ Block only if same student already scanned within SAME grading period date range.
    """
    answer_key = get_object_or_404(AnswerKey, pk=pk, user=request.user)
    institution = getattr(answer_key, "institution", None)

    assessment_title = answer_key.quiz_name

    today = timezone.localdate()
    sys_today = date.today()
    if today != sys_today:
        today = sys_today

    gp_id = request.GET.get("gp") or request.POST.get("gp")
    grading_period = None
    if gp_id and institution:
        grading_period = GradingPeriod.objects.filter(id=gp_id, institution=institution).first()
    if not grading_period and institution:
        grading_period = get_grading_period_for_date(institution, today)

    # mode: save/rescan
    mode = (request.GET.get("mode") or request.POST.get("mode") or "save").strip().lower()
    scan_id = request.GET.get("scan_id") or request.POST.get("rescan_id")
    rescan_scan = None
    if mode == "rescan" and scan_id:
        rescan_scan = get_object_or_404(ScanResult, pk=scan_id, answer_key=answer_key)

    # build answer map/points map
    answer_map, points_map = _build_omr_answer_key(answer_key)
    if not answer_map:
        messages.error(request, "This answer key has no answers set yet.")
        return redirect("answer_key_review", pk=answer_key.pk)

    students_qs = Student.objects.filter(section=answer_key.section).order_by("last_name", "full_name")
    selected_student_id = None

    def format_student_name(stu: Student) -> str:
        given = (getattr(stu, "full_name", "") or "").strip()
        mi = (getattr(stu, "middle_initial", "") or "").strip()
        last = (getattr(stu, "last_name", "") or "").strip()
        if last:
            base = last
            if given: base += f", {given}"
            if mi: base += f" {mi}."
            return base
        return given or "Unknown"

    # helper: redirect back to dashboard and open matrix
    def _redirect_to_matrix(extra_message=None, level="success"):
        url = reverse("user_dashboard")
        if grading_period:
            url += f"?gp={grading_period.id}&open=matrix"
        else:
            url += "?open=matrix"

        if extra_message:
            if level == "error":
                messages.error(request, extra_message)
            else:
                messages.success(request, extra_message)

        return redirect(url)

    if request.method == "POST":
        sheet_image = request.FILES.get("sheet_image")
        rescan_id = request.POST.get("rescan_id")

        if not sheet_image:
            return _redirect_to_matrix("Please capture or upload a bubble sheet image.", level="error")

        existing_scan = None

        # RESCAN MODE
        if mode == "rescan" and rescan_id:
            existing_scan = get_object_or_404(ScanResult, pk=rescan_id, answer_key=answer_key)
            student_name = existing_scan.student_name or "Unknown"

        # SAVE MODE
        else:
            student_id = (request.POST.get("student_id") or "").strip()
            selected_student_id = student_id or None

            if not student_id:
                return _redirect_to_matrix("Please select a student.", level="error")

            student = get_object_or_404(students_qs, pk=student_id)
            student_name = format_student_name(student)

            # DUPLICATE CHECK (ONLY WITHIN SAME GRADING PERIOD)
            dup_qs = ScanResult.objects.filter(
                answer_key=answer_key,
                student_name__iexact=student_name
            )

            if grading_period:
                dup_qs = dup_qs.filter(
                    created_at__date__gte=grading_period.start_date,
                    created_at__date__lte=grading_period.end_date
                )

            dup = dup_qs.first()
            if dup:
                return _redirect_to_matrix(
                    f"{student_name} already has a record for '{assessment_title}' for this grading period. Use Re-scan or delete in history.",
                    level="error"
                )

        # -------------------------
        # SAVE RAW IMAGE FILE
        # -------------------------
        base_dir = os.path.join(
            settings.MEDIA_ROOT,
            "bubble_scans",
            str(institution.id),        # Short ID
            _safe_folder_name(institution.school_year), 
            str(answer_key.grade.id),   # Short ID
            str(answer_key.section.id), # Short ID
            str(answer_key.subject.id), # Short ID
        )
        os.makedirs(base_dir, exist_ok=True)

        raw_name = f"s_{uuid.uuid4().hex[:12]}.jpg"
        raw_path = os.path.join(base_dir, raw_name)
        with open(raw_path, "wb+") as dest:
            for chunk in sheet_image.chunks():
                dest.write(chunk)

        # -------------------------
        # OMR ANALYSIS (KNN)
        # -------------------------
        knn, class_names, img_size, cal = get_knn()
        warped, raw_results, layout_id = analyze_sheet(Path(raw_path), knn, class_names, img_size, cal)
        summary = score_sheet(raw_results, answer_map)


        # build answers to save
        per_item = summary.get("per_item", {})
        answers_to_save = []
        for q_no in sorted(answer_map.keys()):
            s = per_item.get(q_no, {})
            status = s.get("status", "blank")
            student_ans = s.get("student_answer") or ""
            answers_to_save.append("INVALID" if status == "invalid" else student_ans)

        total_points = sum(points_map.values()) if points_map else 0
        score = summary.get("score_raw", 0)
        max_score = total_points

        # annotated image
        annotated = annotate_sheet(warped, summary, layout_id)
        annot_name = f"annotated_{raw_name}"
        annot_full_path = os.path.join(base_dir, annot_name)
        cv2.imwrite(annot_full_path, annotated)

        rel_path = os.path.relpath(annot_full_path, settings.MEDIA_ROOT).replace("\\", "/")

        # -------------------------
        # SAVE/UPDATE ScanResult
        # -------------------------
        if mode == "rescan" and existing_scan:
            existing_scan.student_name = student_name
            existing_scan.sheet_image = rel_path
            existing_scan.answers = answers_to_save
            existing_scan.score = score
            existing_scan.max_score = max_score
            existing_scan.save()
        else:
            ScanResult.objects.create(
                answer_key=answer_key,
                grade=answer_key.grade,
                section=answer_key.section,
                subject=answer_key.subject,
                student_name=student_name,
                sheet_image=rel_path,
                answers=answers_to_save,
                score=score,
                max_score=max_score,
            )

        return _redirect_to_matrix(f"Scan saved for {student_name} — {assessment_title}.")

    # GET (you can keep this or redirect too)
    # Usually you don't want to land here directly anymore,
    # but keep it safe:
    return _redirect_to_matrix()

# ------------------------------------------------------------------------------
# VIEW PREVIOUS SCAN RESULT (FROM HISTORY)
# ------------------------------------------------------------------------------



@login_required
def view_scan_result(request, pk):
    scan = get_object_or_404(ScanResult, pk=pk)

    # Security: only owner of the answer key can view
    if scan.answer_key.user_id != request.user.id:
        return HttpResponse("Unauthorized", status=403)

    answer_key = scan.answer_key

    # correct answers from uploaded answer key file
    correct_answers = read_answer_key_file(answer_key.file)
    num_questions = len(correct_answers)

    # student answers from scan.answers (supports list of strings or list of dicts)
    raw_answers = scan.answers or []
    student_answers = []

    if isinstance(raw_answers, list):
        if raw_answers and isinstance(raw_answers[0], dict):
            # expects something like [{"number":1,"student":"A"}, ...]
            # convert to ordered list
            tmp = {int(a.get("number", 0)): (a.get("student") or "") for a in raw_answers}
            for i in range(1, num_questions + 1):
                student_answers.append(tmp.get(i, ""))
        else:
            # expects something like ["A","B","",...]
            student_answers = [str(x or "") for x in raw_answers]
    else:
        student_answers = []

    # pad / cut to match number of questions
    if len(student_answers) < num_questions:
        student_answers += [""] * (num_questions - len(student_answers))
    if len(student_answers) > num_questions:
        student_answers = student_answers[:num_questions]

    points_dict = (answer_key.item_points or {})
    total_correct = 0
    items = []

    for idx, correct in enumerate(correct_answers, start=1):
        student = (student_answers[idx - 1] or "").strip()
        correct = (correct or "").strip()

        is_correct = bool(student) and student == correct
        if is_correct:
            total_correct += 1

        points = points_dict.get(str(idx), points_dict.get(idx, 1))
        try:
            points = float(points)
        except Exception:
            points = 1

        items.append({
            "number": idx,
            "correct": correct,
            "student": student,
            "is_correct": is_correct,
            "points": points,
        })

    max_score = float(scan.max_score or 0)
    score = float(scan.score or 0)
    score_percent = (score / max_score * 100) if max_score else 0.0

    # compute grading period based on scan date (THIS is the "where grading siya nagscan")
    scan_date = scan.created_at.date() if scan.created_at else timezone.localdate()
    grading_period = get_grading_period_for_date(scan.institution or answer_key.institution, scan_date)

    context = {
        "scan": scan,
        "answer_key": answer_key,
        "grade": scan.grade,
        "section": scan.section,
        "subject": scan.subject,

        "student_name": scan.student_name,
        "items": items,
        "options": ["A", "B", "C", "D", "E"],

        "total_items": num_questions,
        "total_correct": total_correct,
        "score": score,
        "max_score": max_score,
        "score_percent": score_percent,

        "image_url": scan.sheet_image.url if scan.sheet_image else "",
        "grading_period": grading_period,
        "gp_id": grading_period.id if grading_period else "",
    }

    return render(request, "analysis/users/user_dashboard", context)



# ------------------------------------------------------------------------------
# SAVE SNAPSHOT (BASE64) FROM CAMERA
# ------------------------------------------------------------------------------

@csrf_exempt
def save_snapshot(request, pk):
    """
    Receives a base64 PNG from the camera snapshot and saves it as a file.
    Returns JSON with the saved file URL.
    """
    if request.method != "POST":
        return JsonResponse(
            {"status": "error", "message": "POST required"},
            status=400,
        )

    try:
        body = request.body.decode("utf-8")
        data = json.loads(body)
        image_data = data.get("image")

        if not image_data:
            return JsonResponse(
                {"status": "error", "message": "No image data provided"},
                status=400,
            )

        if "," in image_data:
            _, base64_data = image_data.split(",", 1)
        else:
            base64_data = image_data

        image_bytes = base64.b64decode(base64_data)

        capture_dir = os.path.join(settings.MEDIA_ROOT, "captures")
        os.makedirs(capture_dir, exist_ok=True)

        filename = f"capture_{pk}_{uuid.uuid4().hex}.png"
        file_path = os.path.join(capture_dir, filename)

        with open(file_path, "wb") as f:
            f.write(image_bytes)

        file_url = f"{settings.MEDIA_URL}captures/{filename}"

        return JsonResponse({"status": "success", "file": file_url})

    except Exception as e:
        return JsonResponse(
            {"status": "error", "message": str(e)},
            status=500,
        )
    
@login_required
def upload_bubble_image(request, pk):
    """
    Backwards-compatibility wrapper.
    Some older URLs/templates may still call 'upload_bubble_image'.
    We just delegate to the main scan_document view which expects pk.
    """
    return scan_document(request, pk)



@login_required
def edit_section(request, grade_id, section_id):
    institution = get_current_institution(request)
    grade = get_object_or_404(Grade, id=grade_id, institution=institution)
    section = get_object_or_404(Section, id=section_id, grade=grade)

    if request.method == "POST":
        form = SectionForm(request.POST, instance=section)
        if form.is_valid():
            section = form.save()

            # AJAX RESPONSE
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return HttpResponse("OK")

            messages.success(request, "Section updated successfully.")
            return redirect("add_grade")

    return HttpResponse(status=405)

def delete_section(request, grade_id, section_id):
    grade = get_object_or_404(Grade, id=grade_id)
    section = get_object_or_404(Section, id=section_id, grade=grade)

    if request.method == "POST":
        section.delete()
        return redirect("add_section", grade_id=grade.id)

    return render(
        request,
        "analysis/confirm_delete_section.html",
        {"grade": grade, "section": section},
    )


@login_required
def edit_subject(request, grade_id, subject_id):
    institution = get_current_institution(request)
    grade = get_object_or_404(Grade, id=grade_id, institution=institution)
    subject = get_object_or_404(Subject, id=subject_id, grade=grade)

    if request.method == "POST":
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()

            # AJAX RESPONSE
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return HttpResponse("OK")

            messages.success(request, "Subject updated successfully.")
            return redirect("add_grade")

    return HttpResponse(status=405)

@login_required
# @user_passes_test(is_admin)
def delete_subject(request, grade_id, subject_id):
    """
    Delete subject then return to add_subject page.
    """
    grade = get_object_or_404(Grade, id=grade_id)
    subject = get_object_or_404(Subject, id=subject_id, grade=grade)

    if request.method == "POST":
        name = subject.name
        subject.delete()
        messages.success(
            request, f"Subject '{name}' deleted successfully from {grade.name}."
        )
    else:
        messages.error(request, "Invalid request method.")

    return redirect("add_subject", grade_id=grade.id)

@login_required
def delete_scan_result(request, pk):
    scan = get_object_or_404(ScanResult, pk=pk)

    # Security: only owner of answer key can delete
    if scan.answer_key.user_id != request.user.id:
        return HttpResponse("Unauthorized", status=403)

    if request.method == "POST":
        gp_id = (request.POST.get("gp") or "").strip()

        grade_id = scan.grade_id
        section_id = scan.section_id
        subject_id = scan.subject_id

        scan.delete()
        messages.success(request, "Scan deleted successfully.")

        # keep grading tab on redirect
        url = reverse("user_capture", args=[grade_id, section_id, subject_id])
        if gp_id:
            url += f"?gp={gp_id}"
        return redirect(url)

    # if not POST, just go back
    return redirect("user_capture", scan.grade_id, scan.section_id, scan.subject_id)

from django.views.decorators.clickjacking import xframe_options_sameorigin

@login_required
@xframe_options_sameorigin
def item_analysis(request, answer_key_id):
    """
    Item analysis for a specific answer key.
    IMPORTANT: Scans follow the AnswerKey (grading), NOT scan created_at date-range.
    """
    answer_key = get_object_or_404(AnswerKey, pk=answer_key_id, user=request.user)

    institution = get_current_institution(request) or getattr(answer_key, "institution", None)

    gp_id = (request.GET.get("gp") or "").strip()

    grading_period = getattr(answer_key, "grading_period", None)

    if gp_id and institution:
        gp_obj = GradingPeriod.objects.filter(id=gp_id, institution=institution).first()
        if gp_obj:
            grading_period = gp_obj

    scans_qs = ScanResult.objects.filter(answer_key=answer_key).order_by("created_at")
    total_students = scans_qs.count()

    answer_map, points_map = _build_omr_answer_key(answer_key)
    max_q = max(answer_map.keys()) if answer_map else 0

    item_stats = []
    for q_no in range(1, max_q + 1):
        correct_choice = (answer_map.get(q_no) or "").strip().upper()
        if not correct_choice:
            continue

        n_correct = 0
        for scan in scans_qs:
            student_answers = scan.answers or []
            if len(student_answers) >= q_no:
                stud = (student_answers[q_no - 1] or "").strip().upper()
                if stud == correct_choice:
                    n_correct += 1

        percent = (n_correct / total_students * 100.0) if total_students > 0 else 0.0
        is_difficult = percent < 40.0

        item_stats.append({
            "number": q_no,
            "correct_choice": correct_choice,
            "n_correct": n_correct,
            "n_students": total_students,
            "percent": percent,
            "is_difficult": is_difficult,
        })

    context = {
        "answer_key": answer_key,
        "grade": answer_key.grade,
        "section": answer_key.section,
        "subject": answer_key.subject,
        "total_students": total_students,
        "item_stats": item_stats,
        "grading_period": grading_period,
        "gp_id": grading_period.id if grading_period else "",
    }
    return render(request, "analysis/users/item_analysis.html", context)

@login_required
def revise_difficult_items(request, item_id, answer_key_id):
    """
    View to revise a question based on difficulty level.
    This view allows users to revise questions that are tagged as 'Revise or Discard'.
    """
    answer_key = get_object_or_404(AnswerKey, pk=answer_key_id, user=request.user)
    item = get_object_or_404(Item, pk=item_id, answer_key=answer_key)

    if request.method == 'POST':
        form = ItemRevisionForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('item_analysis', answer_key_id=answer_key.id)
    else:
        form = ItemRevisionForm(instance=item)

    return render(request, 'analysis/users/revise_item.html', {'form': form, 'item': item})

def _compute_item_difficulty(answer_key):
    """
    Returns a list of dicts:
      [
        {
          "number": q_no,
          "percent": float,
          "n_correct": int,
          "n_students": int,
        },
        ...
      ]
    based on current ScanResult data and current answer key.
    """

    scans_qs = ScanResult.objects.filter(answer_key=answer_key).order_by("created_at")
    total_students = scans_qs.count()

    answer_map, _points_map = _build_omr_answer_key(answer_key)

    if answer_map:
        max_q = max(answer_map.keys())
    else:
        max_q = 0

def _rescore_all_scans_for_answer_key(answer_key):
    """
    Recompute score and max_score for all ScanResult rows linked
    to this AnswerKey, based on the CURRENT item_points/_answers
    of the AnswerKey and the stored ScanResult.answers.

    - Uses 'INVALID' markers in ScanResult.answers to keep invalid items.
    - Does NOT change ScanResult.answers (student answers stay the same).
    """
    # Build current correct answers + points
    answer_map, points_map = _build_omr_answer_key(answer_key)
    if not answer_map:
        return  # nothing to rescore

    scans = ScanResult.objects.filter(answer_key=answer_key)

    for scan in scans:
        student_answers = list(scan.answers or [])

        total_points = 0.0
        max_points = 0.0

        # We also recompute some counts (not stored, but good to be consistent)
        total_correct = 0
        incorrect_count = 0
        blank_count = 0
        invalid_count = 0

        # loop using current answer map
        for q_no in sorted(answer_map.keys()):
            correct = (answer_map.get(q_no) or "").strip().upper()
            pts = points_map.get(q_no, 1.0)

            # get stored code for this question ("A"/"B"/... or "INVALID" or "")
            if q_no - 1 < len(student_answers):
                raw_code = student_answers[q_no - 1] or ""
            else:
                raw_code = ""
            code = raw_code.strip().upper()

            # classify
            if code == "INVALID":
                status = "invalid"
                is_correct = False
                invalid_count += 1
            else:
                student = code  # letter or ""
                if not student:
                    status = "blank"
                    is_correct = False
                    blank_count += 1
                elif student == correct and correct:
                    status = "correct"
                    is_correct = True
                    total_correct += 1
                else:
                    status = "wrong"
                    is_correct = False
                    incorrect_count += 1

            max_points += pts
            if is_correct:
                total_points += pts

        # update scan row
        scan.score = total_points
        scan.max_score = max_points
        scan.save(update_fields=["score", "max_score"])

def _reannotate_all_scans_for_answer_key(answer_key):
    """
    Re-generate the annotated bubble-sheet image for every ScanResult
    linked to this AnswerKey, using the CURRENT answer key.
    """
    # Build the current answer map (so colors match the new key)
    answer_map, _points_map = _build_omr_answer_key(answer_key)
    if not answer_map:
        return

    scans = ScanResult.objects.filter(answer_key=answer_key)

    for scan in scans:
        if not scan.sheet_image:
            continue

        # Absolute path to the existing annotated image
        annot_rel = str(scan.sheet_image) 
        annot_full = os.path.join(settings.MEDIA_ROOT, annot_rel.replace("/", os.sep))

        base_dir = os.path.dirname(annot_full)
        fname = os.path.basename(annot_full)

        # Infer original raw sheet filename: strip "annotated_"
        if fname.startswith("annotated_"):
            raw_name = fname[len("annotated_"):]
        else:
            continue

        raw_path = os.path.join(base_dir, raw_name)
        if not os.path.exists(raw_path):
            continue

        try:
            # 1. UPDATED: Capture layout_id from analyze_sheet
            # This ensures we know if it's a 10, 50, or 120 item layout
            warped, raw_results, layout_id = analyze_sheet(
                Path(raw_path), KNN_MODEL, KNN_CLASS_NAMES, KNN_IMG_SIZE, KNN_CAL
            )

            # 2. Re-calculate the score summary based on the NEW answer_map
            summary = score_sheet(raw_results, answer_map)

            # 3. UPDATED: Pass layout_id to annotate_sheet
            # This prevents the "shifted circles" bug by using correct template coordinates
            annotated = annotate_sheet(warped, summary, layout_id)

            # 4. Overwrite the existing annotated file
            cv2.imwrite(annot_full, annotated)

        except Exception as e:
            print(f"[reannotate] Failed for ScanResult {scan.pk}: {e}")

from django.http import HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect
from django.views.decorators.http import require_POST


@require_POST
@login_required
def superadmin_delete_admin_user(request, user_id):
    # Super Admin only
    if not request.user.is_superuser:
        return HttpResponseForbidden("Unauthorized")

    u = get_object_or_404(User, id=user_id)

    # Safety: don't delete yourself
    if u.id == request.user.id:
        messages.error(request, "You cannot delete your own account.")
        return redirect("create_user")

    # Safety: don't delete superusers
    if u.is_superuser:
        messages.error(request, "You cannot delete a Super Admin account.")
        return redirect("create_user")

    # Optional: only allow deleting admin accounts (staff)
    if not u.is_staff:
        messages.error(request, "This user is not an admin account.")
        return redirect("create_user")

    u.delete()
    messages.success(request, "Admin user deleted successfully.")

    # stay/refresh create admin user page
    return redirect("create_user")

@require_POST
@login_required
def superadmin_delete_institution(request, institution_id):
    # Super Admin only
    if not request.user.is_superuser:
        messages.error(request, "You are not allowed to delete institutions.")
        return redirect("admin_dashboard")  # change if your dashboard name differs

    inst = get_object_or_404(Institution, id=institution_id)
    name = inst.name
    inst.delete()

    messages.success(request, f"Deleted institution: {name}")
    return redirect("add_institution")  # refresh same page

def _save_teacher_assignments(request, teacher, institution):
    """
    Cleans up all previous assignments and saves the new unique selection.
    """
    active_sy = institution.school_year
    
    # 1. DELETE ALL existing assignments for this teacher at this institution.
    # Removing 'school_year' from the filter ensures past years don't linger.
    TeacherClassAssignment.objects.filter(
        teacher=teacher,
        institution=institution
    ).delete()

    grade_ids = request.POST.getlist("assign_grade[]")
    section_ids = request.POST.getlist("assign_section[]")
    subject_ids = request.POST.getlist("assign_subject[]")

    # 2. Use a set to prevent the duplicates seen in your screenshot
    seen_in_form = set()
    first_assignment = None

    for g, s, sub in zip(grade_ids, section_ids, subject_ids):
        if not (g and s and sub):
            continue
        
        # Create a unique key for this combination
        identifier = (str(g), str(s), str(sub))
        
        # Only save if we haven't already processed this exact combo in this request
        if identifier not in seen_in_form:
            new_a = TeacherClassAssignment.objects.create(
                teacher=teacher,
                grade_id=g,
                section_id=s,
                subject_id=sub,
                institution=institution,
                school_year=active_sy
            )
            
            if first_assignment is None:
                first_assignment = new_a
                
            seen_in_form.add(identifier)

    # 3. Update legacy fields on the Teacher model
    if first_assignment:
        teacher.grade_id = first_assignment.grade_id
        teacher.section_id = first_assignment.section_id
        teacher.subject_id = first_assignment.subject_id
    else:
        teacher.grade = teacher.section = teacher.subject = None
    
    teacher.save()

@login_required
def teacher_students(request, grade_id, section_id, subject_id):
    teacher = get_object_or_404(
        Teacher,
        user=request.user,
    )

    institution = get_current_institution(request)
    academic_year = institution.school_year

    # SECURITY: teacher must be assigned to this class
    is_assigned = TeacherClassAssignment.objects.filter(
        teacher=teacher,
        grade_id=grade_id,
        section_id=section_id,
        subject_id=subject_id,
        institution=institution,
        school_year=academic_year,
    ).exists()

    if not is_assigned:
        return HttpResponse("Unauthorized", status=403)

    # Fetch students added/imported by admin
    students = Student.objects.filter(
        institution=institution,
        school_year=academic_year,
        grade_id=grade_id,
        section_id=section_id,
    ).order_by("last_name", "full_name")

    grade = get_object_or_404(Grade, id=grade_id)
    section = get_object_or_404(Section, id=section_id)
    subject = get_object_or_404(Subject, id=subject_id)

    return render(request, "analysis/users/teacher_students.html", {
        "students": students,
        "grade": grade,
        "section": section,
        "subject": subject,
        "institution_name": institution.name,
        "academic_year": academic_year,
    })

def get_grading_period_for_date(institution, d):
    """
    Returns the grading period where date d belongs (start_date <= d <= end_date),
    ordered 1st->4th.
    """
    if not institution or not d:
        return None

    if hasattr(d, "date"):
        d = d.date()

    order_case = Case(
        When(period="1st Grading", then=1),
        When(period="2nd Grading", then=2),
        When(period="3rd Grading", then=3),
        When(period="4th Grading", then=4),
        default=99,
        output_field=IntegerField(),
    )

    return (
        GradingPeriod.objects
        .filter(institution=institution, start_date__lte=d, end_date__gte=d)
        .annotate(_o=order_case)
        .order_by("_o")
        .first()
    )

GRADING_ORDER = {
    "1st Grading": 1,
    "2nd Grading": 2,
    "3rd Grading": 3,
    "4th Grading": 4,
}

def grading_order_case():
    return Case(
        When(period="1st Grading", then=1),
        When(period="2nd Grading", then=2),
        When(period="3rd Grading", then=3),
        When(period="4th Grading", then=4),
        default=99,
        output_field=IntegerField(),
    )

def get_ordered_grading_periods(institution):
    if not institution:
        return []
    oc = grading_order_case()
    return list(
        GradingPeriod.objects
        .filter(institution=institution)
        .annotate(_o=oc)
        .order_by("_o", "start_date")
    )

def get_selected_gp(request, gp_list, institution):
    gp_id = request.GET.get("gp") or request.POST.get("grading_period")
    if gp_id:
        for gp in gp_list:
            if str(gp.id) == str(gp_id):
                return gp

    # fallback: active today
    today = timezone.localdate()
    for gp in gp_list:
        if gp.start_date <= today <= gp.end_date:
            return gp

    # fallback: first by order
    return gp_list[0] if gp_list else None

def build_gp_buttons(gp_list, selected_gp):
    """
    Past grading: ENABLED (clickable)
    ❌ Future grading (start_date > today): DISABLED
    """
    today = timezone.localdate()
    buttons = []
    for gp in gp_list:
        enabled = not (gp.start_date and gp.start_date > today)  # only future disabled
        buttons.append({
            "id": gp.id,
            "period": gp.period,
            "start_date": gp.start_date,
            "end_date": gp.end_date,
            "enabled": enabled,
            "is_selected": (selected_gp and gp.id == selected_gp.id),
        })
    return buttons

@login_required
def students_view(request):
    teacher = (
        Teacher.objects
        .select_related("institution")
        .filter(user=request.user)
        .first()
    )

    institution = get_current_institution(request) or getattr(teacher, "institution", None)
    if not institution or not teacher:
        messages.error(request, "Teacher account / institution context not found.")
        return redirect("user_dashboard")

    academic_year = getattr(institution, "school_year", None)

    assigned_qs = (
        TeacherClassAssignment.objects
        .select_related("grade", "section")
        .filter(
            teacher=teacher,
            institution=institution,
            school_year=academic_year,
        )
    )

    allowed_section_ids = list(assigned_qs.values_list("section_id", flat=True).distinct())

    if request.method == "POST":
        editing_student_id = request.POST.get("student_id") or None
        instance = None

        if editing_student_id:
            instance = get_object_or_404(
                Student,
                id=editing_student_id,
                section_id__in=allowed_section_ids,
            )

        # Teacher can only add students to their own assigned sections
        section_id = request.POST.get("section")
        if section_id and int(section_id) not in set(allowed_section_ids):
            messages.error(request, "You are not allowed to add students to that section.")
            return redirect("user_dashboard")

        form = StudentForm(request.POST, instance=instance)

        if form.is_valid():
            obj = form.save(commit=False)

            # --- GENDER LOGIC FIX ---
            # Capture the gender from the dropdown added to the template
            gender_val = request.POST.get("gender")
            if gender_val:
                # Check if your model uses 'gender' or 'sex' field name
                if hasattr(obj, "gender"):
                    obj.gender = gender_val
                elif hasattr(obj, "sex"):
                    obj.sex = gender_val

            # Attach institution & school year metadata
            if hasattr(obj, "institution"):
                obj.institution = institution
            if hasattr(obj, "school_year") and academic_year:
                obj.school_year = academic_year
            
            # Ensure the grade is set based on the chosen section
            if hasattr(obj, "grade") and obj.section:
                obj.grade = obj.section.grade

            obj.save()
            messages.success(request, "Student saved successfully.")
            return redirect("user_dashboard")

        # If form is invalid, show errors
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
        
        return redirect("user_dashboard")

    return redirect("user_dashboard")
# helper regex
RE_NUMBERED = re.compile(r"^\s*(\d+)\s*[\.\)\-:]\s*(.+)$")  # 1. Name / 1) Name / 1- Name / 1: Name
RE_ANY_NUMBER_IN_LINE = re.compile(r"(\d+\s*[\.\)\-:]\s*)")  # used to strip anything before first number

def _normalize_header(s: str) -> str:
    return str(s or "").strip().lower().replace(" ", "").replace("_", "")

def _clean_name_cell(raw: str) -> str:
    """
    - removes leading numbering like '1. ' / '12) '
    - trims junk
    """
    if raw is None:
        return ""
    text = str(raw).strip()
    if not text:
        return ""

    m = RE_NUMBERED.match(text)
    if m:
        text = m.group(2).strip()

    # common bullets
    text = re.sub(r"^[•\-\u2022]+\s*", "", text).strip()
    return text

def _parse_name_to_parts(name_text: str):
    """
    Accepts formats like:
      - "LastName, FirstName MI"
      - "FirstName MI LastName"
    Returns: first, last, mi
    """
    name_text = (name_text or "").strip()
    if not name_text:
        return ("", "", "")

    # If "Last, First ..."
    if "," in name_text:
        last_part, rest = [p.strip() for p in name_text.split(",", 1)]
        last = last_part.strip()
        parts = rest.split()
        mi = ""
        first = ""

        if parts:
            # detect MI like "A." or "A"
            if len(parts[-1]) <= 3 and parts[-1].rstrip(".").isalpha():
                mi = parts[-1].rstrip(".")
                first = " ".join(parts[:-1]).strip()
            else:
                first = " ".join(parts).strip()

        return (first, last, mi)

    # Else "First ... Last"
    parts = name_text.split()
    if len(parts) >= 2:
        last = parts[-1].strip()
        first = " ".join(parts[:-1]).strip()
        return (first, last, "")

    # single token fallback
    return (name_text, "", "")

def _student_gender_field_name():
    """
    Supports either Student.gender or Student.sex if you used a different field name.
    """
    # NOTE: on Django model class, the attribute exists if the field exists
    from .models import Student
    if hasattr(Student, "gender"):
        return "gender"
    if hasattr(Student, "sex"):
        return "sex"
    return None

@login_required
def import_students_view(request):
    if request.method != "POST":
        return redirect("user_dashboard")

    teacher = Teacher.objects.filter(user=request.user).first()
    institution = get_current_institution(request) or getattr(teacher, "institution", None)

    if not institution:
        messages.error(request, "No institution context found.")
        return redirect("user_dashboard")

    academic_year = getattr(institution, "school_year", None)

    assigned_qs = TeacherClassAssignment.objects.filter(
        teacher=teacher,
        institution=institution,
        school_year=academic_year,
    )
    allowed_section_ids = set(assigned_qs.values_list("section_id", flat=True))

    grade_id = request.POST.get("grade")
    section_id = request.POST.get("section")
    uploaded_file = request.FILES.get("student_file")

    if not (grade_id and section_id and uploaded_file):
        messages.error(request, "Please select a grade, section, and file.")
        return redirect("user_dashboard")

    if int(section_id) not in allowed_section_ids:
        messages.error(request, "You are not allowed to import into that section.")
        return redirect("user_dashboard")

    grade = get_object_or_404(Grade, pk=grade_id, institution=institution)
    section = get_object_or_404(Section, pk=section_id, grade=grade)

    filename = uploaded_file.name.lower()
    created_count = 0
    updated_count = 0

    gender_field = _student_gender_field_name()

    def upsert_student(first, last, mi, gender_value=None):
        nonlocal created_count, updated_count

        first = (first or "").strip()
        last = (last or "").strip()
        mi = (mi or "").strip()

        if not first and not last:
            return

        defaults = {"grade": grade}
        student, created = Student.objects.get_or_create(
            institution=institution,
            school_year=institution.school_year,
            section=section,
            full_name=first,
            last_name=last,
            middle_initial=mi,
            defaults=defaults,
        )

        # update gender if field exists and we got a value
        if gender_field and gender_value:
            current = getattr(student, gender_field, None)
            if not current:
                setattr(student, gender_field, gender_value)
                student.save(update_fields=[gender_field])
                if not created:
                    updated_count += 1

        if created:
            created_count += 1

    try:
        
        # EXCEL (supports Male/Female columns)
        
        if filename.endswith((".xlsx", ".xls")):
            wb = load_workbook(uploaded_file, read_only=True, data_only=True)
            ws = wb.active

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header = [_normalize_header(c) for c in header_row]

            def col_idx(*names):
                for nm in names:
                    nm = _normalize_header(nm)
                    if nm in header:
                        return header.index(nm)
                return None

            col_last = col_idx("lastname", "last name")
            col_first = col_idx("firstnames", "firstname", "first name", "name")
            col_mi = col_idx("middleinitial", "mi", "middle initial")

            col_male = col_idx("male", "boys")
            col_female = col_idx("female", "girls")

            # CASE A: file has MALE/FEMALE columns (one cell per student per row)
            if col_male is not None or col_female is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue

                    cells = list(row)

                    def safe_cell(idx):
                        if idx is None or idx >= len(cells):
                            return ""
                        return _clean_name_cell(cells[idx])

                    male_name = safe_cell(col_male)
                    female_name = safe_cell(col_female)

                    if male_name:
                        first, last, mi = _parse_name_to_parts(male_name)
                        upsert_student(first, last, mi, gender_value="Male")

                    if female_name:
                        first, last, mi = _parse_name_to_parts(female_name)
                        upsert_student(first, last, mi, gender_value="Female")

            # CASE B: classic columns First/Last/MI (no gender column)
            else:
                if col_first is None:
                    messages.error(request, "Excel must have a 'FirstNames' / 'FirstName' / 'Name' column.")
                    return redirect("user_dashboard")

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    cells = list(row)

                    def val(idx):
                        if idx is None or idx >= len(cells):
                            return ""
                        return str(cells[idx]).strip() if cells[idx] is not None else ""

                    first = val(col_first)
                    last = val(col_last)
                    mi = val(col_mi)
                    upsert_student(first, last, mi, gender_value=None)

        
        # CSV (supports Male/Female columns too)
        
        elif filename.endswith(".csv"):
            decoded = uploaded_file.read().decode("utf-8", errors="ignore").splitlines()
            reader = csv.DictReader(decoded)

            # normalized headers map
            norm_map = {_normalize_header(k): k for k in (reader.fieldnames or [])}

            def get_field(row, *names):
                for nm in names:
                    key = norm_map.get(_normalize_header(nm))
                    if key:
                        return row.get(key)
                return None

            has_male = any(_normalize_header(h) in ("male", "boys") for h in (reader.fieldnames or []))
            has_female = any(_normalize_header(h) in ("female", "girls") for h in (reader.fieldnames or []))

            for row in reader:
                if not row:
                    continue

                if has_male or has_female:
                    male_name = _clean_name_cell(get_field(row, "Male", "Boys"))
                    female_name = _clean_name_cell(get_field(row, "Female", "Girls"))

                    if male_name:
                        first, last, mi = _parse_name_to_parts(male_name)
                        upsert_student(first, last, mi, gender_value="Male")

                    if female_name:
                        first, last, mi = _parse_name_to_parts(female_name)
                        upsert_student(first, last, mi, gender_value="Female")
                else:
                    first = (get_field(row, "FirstNames", "FirstName", "Name") or "").strip()
                    last = (get_field(row, "LastName") or "").strip()
                    mi = (get_field(row, "MiddleInitial", "MI") or "").strip()
                    upsert_student(first, last, mi, gender_value=None)

        
        # PDF (detect MALE/FEMALE headers + ignore lines before numbering)
        
        elif filename.endswith(".pdf"):
            reader = PdfReader(uploaded_file)
            lines = []
            for page in reader.pages:
                text = page.extract_text() or ""
                for ln in text.splitlines():
                    ln = (ln or "").strip()
                    if ln:
                        lines.append(ln)

            current_gender = None  # "Male" or "Female"

            for raw_line in lines:
                line = raw_line.strip()
                low = line.lower()

                # Detect headers (your guide)
                if re.fullmatch(r"(male|males|boys)\s*", low):
                    current_gender = "Male"
                    continue
                if re.fullmatch(r"(female|females|girls)\s*", low):
                    current_gender = "Female"
                    continue

                # Ignore anything before the number:
                # - if a line contains a numbered item after some text, keep only from number onwards
                # Example: "Grade 7 MALE 1. Dela Cruz, Juan" -> becomes "1. Dela Cruz, Juan"
                if RE_ANY_NUMBER_IN_LINE.search(line) and not RE_NUMBERED.match(line):
                    # keep substring from first number pattern
                    idx = RE_ANY_NUMBER_IN_LINE.search(line).start()
                    line = line[idx:].strip()

                # Only accept numbered entries
                m = RE_NUMBERED.match(line)
                if not m:
                    continue

                if not current_gender:
                    # If you REQUIRE header to decide gender, skip when unknown
                    continue

                name_text = m.group(2).strip()
                name_text = _clean_name_cell(name_text)

                first, last, mi = _parse_name_to_parts(name_text)
                upsert_student(first, last, mi, gender_value=current_gender)

        else:
            messages.error(request, "Unsupported file type. Upload Excel/CSV/PDF only.")
            return redirect("user_dashboard")

    except Exception as e:
        messages.error(request, f"Could not import students: {e}")
        return redirect("user_dashboard")

    if created_count or updated_count:
        msg = f"Imported {created_count} student(s)"
        if updated_count:
            msg += f" and updated gender for {updated_count} student(s)"
        msg += f" into {grade.name} - {section.name}."
        messages.success(request, msg)
    else:
        messages.info(request, "No new students were imported (they may already exist).")

    return redirect("user_dashboard")

@login_required
@require_POST
def delete_student(request, student_id):
    teacher = Teacher.objects.filter(user=request.user).first()
    institution = get_current_institution(request) or getattr(teacher, "institution", None)

    if not institution or not teacher:
        return redirect("user_dashboard")

    academic_year = getattr(institution, "school_year", None)

    assigned_qs = TeacherClassAssignment.objects.filter(
        teacher=teacher,
        institution=institution,
        school_year=academic_year,
    )
    allowed_section_ids = list(assigned_qs.values_list("section_id", flat=True).distinct())

    student = get_object_or_404(Student, id=student_id, section_id__in=allowed_section_ids)
    student.delete()

    return redirect("user_dashboard")

def _safe_filename(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[^\w\-. ]+", "", s)
    s = s.replace(" ", "_")
    return s[:120] or "file"

def _student_display_name(student) -> str:
    last = (getattr(student, "last_name", "") or "").strip()
    first = (getattr(student, "full_name", "") or "").strip()
    mi = (getattr(student, "middle_initial", "") or "").strip()

    if last:
        name = f"{last}, {first}".strip()
        if mi:
            name += f" {mi}."
        return name
    return first

def _make_overlay_two(page_w: float, page_h: float,
                      top_name: str, top_section: str,
                      bottom_name: str = "", bottom_section: str = "") -> PdfReader:
    packet = BytesIO()
    c = canvas.Canvas(packet, pagesize=(page_w, page_h))
    c.setFont("Helvetica", 11)

    NAME_X = 165
    SECTION_X = 375
    TOP_Y = page_h - 93
    BOTTOM_Y = (page_h / 2) - 66

    if top_name:
        c.drawString(NAME_X, TOP_Y, top_name)
    if top_section:
        c.drawString(SECTION_X, TOP_Y, top_section)

    if bottom_name:
        c.drawString(NAME_X, BOTTOM_Y, bottom_name)
    if bottom_section:
        c.drawString(SECTION_X, BOTTOM_Y, bottom_section)

    c.save()
    packet.seek(0)
    return PdfReader(packet)

@login_required
def download_bubble_sheets_section(request, section_id: int):
    section = get_object_or_404(Section, pk=section_id)

    students = Student.objects.filter(section=section).order_by("last_name", "full_name", "id")
    if not students.exists():
        raise Http404("No students found in this section.")

    template_path = os.path.join(settings.BASE_DIR, "templates_pdf", "Bubble test sheets2.pdf")
    if not os.path.exists(template_path):
        raise Http404(f"Bubble template not found: {template_path}")

    template_reader = PdfReader(template_path)
    if not template_reader.pages:
        raise Http404("Bubble template PDF has no pages.")

    base_page = template_reader.pages[0]
    page_w = float(base_page.mediabox.width)
    page_h = float(base_page.mediabox.height)

    writer = PdfWriter()

    st_list = list(students)
    for i in range(0, len(st_list), 2):
        s1 = st_list[i]
        s2 = st_list[i + 1] if (i + 1) < len(st_list) else None

        overlay_reader = _make_overlay_two(
            page_w, page_h,
            _student_display_name(s1), section.name,
            _student_display_name(s2) if s2 else "", section.name if s2 else ""
        )
        overlay_page = overlay_reader.pages[0]

        # Copy the template page fresh each time, then merge overlay
        page = copy.copy(base_page)
        page.merge_page(overlay_page)
        writer.add_page(page)

    out = BytesIO()
    writer.write(out)
    out.seek(0)

    filename = _safe_filename(f"BubbleSheets_{section.name}.pdf")
    resp = HttpResponse(out.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

def _build_bubbles_pdf(section, students_qs, per_page: int) -> bytes:
    template_map = {
        1: "Bubble test sheets.pdf",   # 1 bubble sheet per page
        2: "Bubble test sheets2.pdf",  # 2 bubble sheets per page
    }
    template_path = os.path.join(settings.BASE_DIR, "templates_pdf", template_map[per_page])
    if not os.path.exists(template_path):
        raise Http404(f"Bubble template not found: {template_path}")

    template_reader = PdfReader(template_path)
    if not template_reader.pages:
        raise Http404("Bubble template PDF has no pages.")

    base_page = template_reader.pages[0]
    page_w = float(base_page.mediabox.width)
    page_h = float(base_page.mediabox.height)

    writer = PdfWriter()
    st_list = list(students_qs)

    if per_page == 1:
        for s in st_list:
            overlay = _make_overlay_one(page_w, page_h, _student_display_name(s), section.name).pages[0]
            page = copy.copy(base_page)
            page.merge_page(overlay)
            writer.add_page(page)

    else:  # per_page == 2
        for i in range(0, len(st_list), 2):
            s1 = st_list[i]
            s2 = st_list[i + 1] if i + 1 < len(st_list) else None

            overlay = _make_overlay_two(
                page_w, page_h,
                _student_display_name(s1), section.name,
                _student_display_name(s2) if s2 else "", section.name if s2 else ""
            ).pages[0]

            page = copy.copy(base_page)
            page.merge_page(overlay)
            writer.add_page(page)

    out = BytesIO()
    writer.write(out)
    return out.getvalue()


@login_required
def download_bubble_sheets_selected(request, section_id: int):
    """
    Downloads ONE PDF for SELECTED students only.
    POST:
      student_ids=1&student_ids=2...
      per_page=1|2
    """
    if request.method != "POST":
        return JsonResponse({"error": "POST required."}, status=405)

    section = get_object_or_404(Section, pk=section_id)

    per_page = (request.POST.get("per_page") or "2").strip()
    if per_page not in ("1", "2"):
        per_page = "2"
    per_page = int(per_page)

    ids = request.POST.getlist("student_ids")
    ids = [int(x) for x in ids if str(x).isdigit()]
    if not ids:
        return JsonResponse({"error": "No students selected."}, status=400)

    students_qs = Student.objects.filter(section=section, id__in=ids).order_by("last_name", "full_name", "id")
    if not students_qs.exists():
        return JsonResponse({"error": "Selected students not found in this section."}, status=404)

    pdf_bytes = _build_bubbles_pdf(section, students_qs, per_page)

    filename = _safe_filename(f"BubbleSheets_{section.name}_selected_{per_page}_per_page.pdf")
    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

@login_required
@require_POST
def delete_all_students_section(request, section_id):
    teacher = Teacher.objects.filter(user=request.user).first()
    institution = get_current_institution(request) or getattr(teacher, "institution", None)
    
    if not institution or not teacher:
        return redirect("user_dashboard")

    academic_year = getattr(institution, "school_year", None)
    
    # Security: Verify the teacher is assigned to this section before deleting
    is_assigned = TeacherClassAssignment.objects.filter(
        teacher=teacher,
        institution=institution,
        section_id=section_id,
        school_year=academic_year
    ).exists()

    if not is_assigned:
        messages.error(request, "Unauthorized action.")
        return redirect("user_dashboard")

    # Perform deletion
    deleted_count, _ = Student.objects.filter(
        section_id=section_id, 
        institution=institution,
        school_year=academic_year
    ).delete()

    messages.success(request, f"Successfully deleted all {deleted_count} students from this section.")
    return redirect("user_dashboard")

# if num_items <= 10: template_file = "1-10 items.pdf"
#     elif num_items <= 15: template_file = "1-15items.pdf"
#     elif num_items <= 20: template_file = "1-20items.pdf"
#     elif num_items <= 25: template_file = "1-25items.pdf"
#     elif num_items <= 30: template_file = "1-30items.pdf"
#     elif num_items <= 35: template_file = "1-35items.pdf"
#     elif num_items <= 40: template_file = "1-40items.pdf"
#     elif num_items <= 45: template_file = "1-45items.pdf"
#     elif num_items <= 50: template_file = "1-50items.pdf"
#     elif num_items <= 55: template_file = "1-55items.pdf"
#     elif num_items <= 60: template_file = "1-60items.pdf"
#     elif num_items <= 65: template_file = "1-65items.pdf"
#     elif num_items <= 70: template_file = "1-70items.pdf"
#     elif num_items <= 80: template_file = "1-80items.pdf"
#     elif num_items <= 90: template_file = "1-90items.pdf"
#     elif num_items <= 100: template_file = "1-100items.pdf"
#     elif num_items <= 110: template_file = "1-110items.pdf"
#     else: template_file = "1-120items.pdf"
